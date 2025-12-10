"""
File Utils MCP Server - 通用文件处理工具

提供通用的文件处理功能：
1. 从 HTTP/HTTPS URL 下载文件
2. 将 CSV/Excel 文件转换为 FASTA 格式
"""

from mcp.server.fastmcp import FastMCP
from starlette.requests import Request
from starlette.responses import StreamingResponse
import tempfile
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Union, Tuple, Optional, AsyncIterator, TYPE_CHECKING, Sequence
import uuid
import time
import os
import urllib.request
import asyncio
import threading
from pydantic import BaseModel, Field
from urllib.parse import urlparse
try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
import logging
from contextlib import asynccontextmanager

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger('FileUtils_MCP')

# Create MCP server
mcp = FastMCP("File Utils Server")

# 全局任务状态存储（用于自定义 SSE 端点）
_task_streams: dict[str, asyncio.Queue] = {}
_task_streams_lock = threading.Lock()

# 添加自定义 SSE 流式传输端点
@mcp.custom_route("/stream/{task_id}", methods=["GET"])
async def stream_task_progress(request: Request):
    """
    自定义 SSE 端点，用于流式传输任务进度
    
    客户端通过 GET /stream/{task_id} 连接，服务器会通过 SSE 推送进度消息
    """
    from starlette.responses import StreamingResponse
    import json
    
    # 从路径参数获取 task_id
    task_id = request.path_params.get("task_id")
    if not task_id:
        from starlette.responses import JSONResponse
        return JSONResponse(
            {"error": "task_id is required in path"},
            status_code=400
        )
    
    # 检查任务是否存在
    with _task_streams_lock:
        if task_id not in _task_streams:
            from starlette.responses import JSONResponse
            return JSONResponse(
                {"error": f"Task {task_id} not found"},
                status_code=404
            )
        queue = _task_streams[task_id]
    
    async def event_generator():
        """SSE 事件生成器"""
        try:
            while True:
                try:
                    # 从队列获取消息（超时 1 秒）
                    message = await asyncio.wait_for(queue.get(), timeout=1.0)
                    
                    # 检查是否是结束标记
                    if message is None:
                        # 发送结束事件
                        yield f"data: {json.dumps({'type': 'end'})}\n\n"
                        break
                    
                    # 发送消息
                    yield f"data: {json.dumps(message, ensure_ascii=False)}\n\n"
                    
                except asyncio.TimeoutError:
                    # 发送心跳，保持连接
                    yield ": heartbeat\n\n"
                    continue
        except Exception as e:
            logger.error(f"[SSE流式传输] 错误 (任务: {task_id}): {str(e)}", exc_info=True)
            yield f"data: {json.dumps({'type': 'error', 'message': str(e)})}\n\n"
        finally:
            # 清理任务队列
            with _task_streams_lock:
                if task_id in _task_streams:
                    del _task_streams[task_id]
    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )

# 实现真正的流式传输：通过猴子补丁修改底层服务器的 handler
# 关键：底层服务器在 server.py:496 调用 list(unstructured_content)，这会立即消费整个生成器
# 解决方案：修改底层服务器的 handler，让它能够处理生成器并实现流式传输
import inspect
from typing import Any
from mcp.types import TextContent, CallToolRequest, ServerResult, CallToolResult
import json

# 修改底层服务器的 handler 来支持异步生成器的流式传输
# 确保 _mcp_server 已初始化
if hasattr(mcp, '_mcp_server') and CallToolRequest in mcp._mcp_server.request_handlers:
    _original_handler = mcp._mcp_server.request_handlers[CallToolRequest]
    
    # Monkey patch RequestResponder.respond 来追踪响应发送
    from mcp.shared.session import RequestResponder
    _original_respond = RequestResponder.respond
    
    async def _patched_respond(self, response):
        """追踪响应发送的 monkey patch"""
        request_id = getattr(self, 'request_id', 'unknown')
        respond_start_time = time.time()
        logger.info(f"[流式传输] 🔵 RequestResponder.respond 被调用 (request_id: {request_id}, response_type: {type(response)})")
        try:
            if hasattr(response, 'model_dump'):
                response_dict = response.model_dump()
                logger.info(f"[流式传输] 🔵 响应内容预览: {json.dumps(response_dict, ensure_ascii=False, default=str)[:500]}...")
            else:
                logger.info(f"[流式传输] 🔵 响应内容: {str(response)[:500]}...")
        except Exception as e:
            logger.warning(f"[流式传输] 🔵 无法序列化响应: {str(e)}")
        
        # 调用原始方法
        logger.info(f"[流式传输] 🔵 准备调用 _original_respond (request_id: {request_id})")
        result = await _original_respond(self, response)
        respond_elapsed = time.time() - respond_start_time
        logger.info(f"[流式传输] ✅ RequestResponder.respond 完成 (request_id: {request_id}, 耗时: {respond_elapsed:.3f}秒)")
        if respond_elapsed > 0.1:
            logger.warning(f"[流式传输] ⚠️ RequestResponder.respond 耗时过长 (request_id: {request_id}, 耗时: {respond_elapsed:.3f}秒)，可能被阻塞")
        return result
    
    RequestResponder.respond = _patched_respond
    logger.info("[流式传输] ✅ 已安装 RequestResponder.respond monkey patch")
    
    # Monkey patch _send_response 来追踪消息发送
    from mcp.shared.session import BaseSession
    _original_send_response = BaseSession._send_response
    
    async def _patched_send_response(self, request_id, response):
        """追踪消息发送的 monkey patch，尝试非阻塞发送"""
        send_start_time = time.time()
        logger.info(f"[流式传输] 🔵 BaseSession._send_response 被调用 (request_id: {request_id}, response_type: {type(response)})")
        
        # 尝试非阻塞发送：使用 asyncio.create_task 在后台发送
        # 但这样会导致响应可能无法立即发送，所以还是需要等待
        # 关键问题：_write_stream.send 可能被阻塞，因为接收端没有及时读取
        
        # 检查是否是 ServerResult 类型，如果是，尝试立即发送
        try:
            # 调用原始方法，但添加超时保护
            import asyncio
            try:
                result = await asyncio.wait_for(
                    _original_send_response(self, request_id, response),
                    timeout=1.0  # 1秒超时
                )
                send_elapsed = time.time() - send_start_time
                logger.info(f"[流式传输] ✅ BaseSession._send_response 完成 (request_id: {request_id}, 耗时: {send_elapsed:.3f}秒)")
                if send_elapsed > 0.1:
                    logger.warning(f"[流式传输] ⚠️ BaseSession._send_response 耗时过长 (request_id: {request_id}, 耗时: {send_elapsed:.3f}秒)，可能被阻塞")
                return result
            except asyncio.TimeoutError:
                logger.error(f"[流式传输] ❌ BaseSession._send_response 超时 (request_id: {request_id})，_write_stream.send 被阻塞超过1秒")
                # 即使超时，也继续等待，因为响应必须发送
                result = await _original_send_response(self, request_id, response)
                send_elapsed = time.time() - send_start_time
                logger.info(f"[流式传输] ✅ BaseSession._send_response 最终完成 (request_id: {request_id}, 总耗时: {send_elapsed:.3f}秒)")
                return result
        except Exception as e:
            logger.error(f"[流式传输] ❌ BaseSession._send_response 异常 (request_id: {request_id}): {str(e)}", exc_info=True)
            raise
    
    # 需要找到实际的 session 实例来 patch
    # 由于 BaseSession 是基类，我们需要在运行时 patch 实例方法
    # 但更好的方法是在 ServerSession 中 patch
    try:
        from mcp.server.lowlevel.server import ServerSession
        ServerSession._send_response = _patched_send_response
        logger.info("[流式传输] ✅ 已安装 ServerSession._send_response monkey patch")
    except Exception as e:
        logger.warning(f"[流式传输] ⚠️ 无法安装 ServerSession._send_response monkey patch: {str(e)}")
    
    # Monkey patch _handle_request 来追踪响应传递
    # 注意：从实例获取的方法已经是绑定的，调用时不需要传递 self
    _original_handle_request = mcp._mcp_server._handle_request
    
    async def _patched_handle_request(self, message, req, session, lifespan_context, raise_exceptions):
        """追踪请求处理的 monkey patch"""
        request_id = getattr(message, 'request_id', 'unknown')
        logger.info(f"[流式传输] 🔵 _handle_request 开始 (request_id: {request_id}, request_type: {type(req).__name__})")
        
        # 调用原始方法（_original_handle_request 已经是绑定方法，不需要传递 self）
        try:
            result = await _original_handle_request(message, req, session, lifespan_context, raise_exceptions)
            logger.info(f"[流式传输] ✅ _handle_request 完成 (request_id: {request_id})")
            return result
        except Exception as e:
            logger.error(f"[流式传输] ❌ _handle_request 异常 (request_id: {request_id}): {str(e)}", exc_info=True)
            raise
    
    # 直接替换方法，Python 会自动处理绑定
    import types
    mcp._mcp_server._handle_request = types.MethodType(_patched_handle_request, mcp._mcp_server)
    logger.info("[流式传输] ✅ 已安装 _handle_request monkey patch")
    
    # 创建新的 handler，它能够处理异步生成器并实现真正的流式传输
    async def _streaming_handler(req: CallToolRequest):
        """支持异步生成器的 handler，实现真正的流式传输"""
        import jsonschema
        
        handler_start_time = time.time()
        logger.info(f"[流式传输] Handler 开始处理请求 (工具: {req.params.name if hasattr(req, 'params') else 'unknown'}, 请求ID: {req.id if hasattr(req, 'id') else 'unknown'})")
        
        try:
            tool_name = req.params.name
            arguments = req.params.arguments or {}
            tool = await mcp._mcp_server._get_cached_tool_definition(tool_name)
            
            # input validation
            if tool:
                try:
                    jsonschema.validate(instance=arguments, schema=tool.inputSchema)
                except jsonschema.ValidationError as e:
                    return mcp._mcp_server._make_error_result(f"Input validation error: {e.message}")
            
            # tool call - 使用原始的 call_tool，但不转换结果
            context = mcp.get_context()
            results = await mcp._tool_manager.call_tool(tool_name, arguments, context=context, convert_result=False)
            
            logger.info(f"[流式传输] 工具调用完成 (工具: {tool_name}), 返回类型: {type(results)}, 是否为异步生成器: {inspect.isasyncgen(results)}")
            
            # 检查是否是异步生成器
            if inspect.isasyncgen(results):
                logger.info(f"[流式传输] 检测到异步生成器 (工具: {tool_name})")
                
                # 方案：使用自定义 SSE 端点实现真正的流式传输
                # 1. 生成任务 ID
                task_id = str(uuid.uuid4())
                
                # 2. 创建消息队列
                message_queue = asyncio.Queue()
                with _task_streams_lock:
                    _task_streams[task_id] = message_queue
                
                logger.info(f"[流式传输] 创建任务流 (工具: {tool_name}, 任务ID: {task_id})")
                
                # 3. 在后台任务中收集消息并推送到队列
                async def collect_and_push():
                    try:
                        logger.info(f"[流式传输] 开始收集异步生成器的消息 (工具: {tool_name}, 任务ID: {task_id})")
                        async for item in results:
                            # 转换每个 item 为字典
                            if isinstance(item, dict):
                                await message_queue.put(item)
                                logger.info(f"[流式传输] 推送消息到队列 (工具: {tool_name}, 任务ID: {task_id}): {json.dumps(item, ensure_ascii=False)[:200]}...")
                            else:
                                await message_queue.put({"type": "data", "content": str(item)})
                        # 发送结束标记
                        await message_queue.put(None)
                        logger.info(f"[流式传输] 消息收集完成 (工具: {tool_name}, 任务ID: {task_id})")
                    except Exception as e:
                        logger.error(f"[流式传输] 收集消息时出错 (工具: {tool_name}, 任务ID: {task_id}): {str(e)}", exc_info=True)
                        await message_queue.put({"type": "error", "message": str(e)})
                        await message_queue.put(None)
                
                # 4. 在新线程中启动后台任务，避免占用当前事件循环
                # 关键：使用独立线程运行事件循环，确保不会阻塞响应发送
                import threading
                def run_in_thread():
                    """在新线程中运行事件循环来执行收集任务"""
                    new_loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(new_loop)
                    try:
                        logger.info(f"[流式传输] 在新线程中启动事件循环 (工具: {tool_name}, 任务ID: {task_id})")
                        new_loop.run_until_complete(collect_and_push())
                    finally:
                        new_loop.close()
                
                thread = threading.Thread(target=run_in_thread, daemon=True)
                thread.start()
                logger.info(f"[流式传输] 已在新线程中启动收集任务 (工具: {tool_name}, 任务ID: {task_id})")
                
                # 5. 返回任务 ID 和服务 ID
                service_id = "file_utils"
                
                logger.info(f"[流式传输] 返回任务信息 (工具: {tool_name}, 任务ID: {task_id}, 服务ID: {service_id})")
                
                # 立即返回任务 ID 和服务 ID
                try:
                    result = ServerResult(
                        CallToolResult(
                            content=[TextContent(
                                type="text",
                                text=json.dumps({
                                    "type": "streaming_task",
                                    "task_id": task_id,
                                    "service_id": service_id,
                                    "message": "任务已启动，请通过 service_id 和 task_id 构建 SSE 端点 URL 来接收实时进度消息。URL 格式: http://{host}:{port}/stream/{task_id}，其中 host 和 port 由客户端根据 service_id 确定。"
                                }, ensure_ascii=False)
                            )],
                            structuredContent={
                                "type": "streaming_task",
                                "task_id": task_id,
                                "service_id": service_id,
                                "message": "任务已启动，请通过 service_id 和 task_id 构建 SSE 端点 URL"
                            },
                            isError=False,
                        )
                    )
                    
                    logger.info(f"[流式传输] 准备返回结果 (工具: {tool_name}, 任务ID: {task_id})")
                    
                    # 尝试序列化结果以确认格式正确
                    try:
                        if hasattr(result, 'model_dump'):
                            result_dict = result.model_dump()
                            logger.info(f"[流式传输] 返回结果类型: ServerResult, 有 model_dump 方法")
                            logger.debug(f"[流式传输] 返回结果内容: {json.dumps(result_dict, ensure_ascii=False, default=str)[:500]}...")
                        else:
                            logger.info(f"[流式传输] 返回结果类型: {type(result)}, 无 model_dump 方法")
                            logger.debug(f"[流式传输] 返回结果内容: {str(result)[:500]}...")
                    except Exception as e:
                        logger.warning(f"[流式传输] 无法序列化返回结果: {str(e)}")
                    
                    handler_elapsed = time.time() - handler_start_time
                    logger.info(f"[流式传输] 正在返回 ServerResult (工具: {tool_name}, 任务ID: {task_id}, Handler耗时: {handler_elapsed:.3f}秒)")
                    logger.info(f"[流式传输] ✅ Handler 返回结果，应该立即发送给客户端 (工具: {tool_name}, 任务ID: {task_id})")
                    
                    # 验证返回值的类型和结构
                    logger.info(f"[流式传输] 返回值类型: {type(result)}")
                    if hasattr(result, 'result'):
                        logger.info(f"[流式传输] result.result 类型: {type(result.result)}")
                        if hasattr(result.result, 'content'):
                            content_type = type(result.result.content)
                            content_len = len(result.result.content) if hasattr(result.result.content, '__len__') else 'N/A'
                            logger.info(f"[流式传输] content 类型: {content_type}, 长度: {content_len}")
                            # 打印 content 的详细信息
                            if isinstance(result.result.content, list) and len(result.result.content) > 0:
                                first_item = result.result.content[0]
                                logger.info(f"[流式传输] content[0] 类型: {type(first_item)}")
                                if hasattr(first_item, 'text'):
                                    logger.info(f"[流式传输] content[0].text 预览: {first_item.text[:200]}...")
                        if hasattr(result.result, 'structuredContent'):
                            logger.info(f"[流式传输] structuredContent: {json.dumps(result.result.structuredContent, ensure_ascii=False)}")
                    
                    logger.info(f"[流式传输] 🔄 准备 return result，FastMCP 应该会处理这个返回值并发送给客户端")
                    
                    # 额外验证：确保 result 可以被序列化
                    try:
                        # 尝试序列化为 JSON 以验证格式
                        if hasattr(result, 'model_dump'):
                            result_json = result.model_dump()
                            json_str = json.dumps(result_json, ensure_ascii=False, default=str)
                            logger.info(f"[流式传输] ✅ ServerResult 可以成功序列化为 JSON，长度: {len(json_str)} 字符")
                        else:
                            logger.warning(f"[流式传输] ⚠️ ServerResult 没有 model_dump 方法")
                    except Exception as e:
                        logger.error(f"[流式传输] ❌ ServerResult 序列化失败: {str(e)}", exc_info=True)
                    
                    return result
                except Exception as e:
                    logger.error(f"[流式传输] 创建返回结果时出错 (工具: {tool_name}, 任务ID: {task_id}): {str(e)}", exc_info=True)
                    return mcp._mcp_server._make_error_result(f"Failed to create streaming task result: {str(e)}")
                
                # 旧的实现（保留作为备选）
                # 实现真正的流式传输：逐个发送多个 ServerResult
                # 关键：我们需要修改底层服务器的代码，让它能够处理生成器
                # 但 handler 的返回值会被底层服务器处理，它期望返回单个 ServerResult
                #
                # 解决方案：返回一个生成器，它会在被迭代时逐个 yield ServerResult
                # 然后修改底层服务器的代码，让它能够处理生成器
                #
                # 但问题是，handler 的返回值会被底层服务器处理，它期望返回单个 ServerResult
                # 所以我们需要修改底层服务器的代码，让它能够处理生成器
                #
                # 或者，我们可以通过其他方式实现流式传输：
                # 1. 使用 MCP 的进度消息机制（如果存在）
                # 2. 或者直接通过 SSE 连接发送消息（需要访问底层传输层）
                #
                # 实际上，FastMCP 的 SSE 传输层可能已经支持流式传输
                # 但需要通过特殊的方式实现
                #
                # 最终方案：创建一个特殊的生成器，它会在被迭代时逐个发送 ServerResult
                # 但底层服务器的 handler 返回的是单个 ServerResult，不是生成器
                # 所以我们需要修改 handler，让它返回一个生成器，而不是单个 ServerResult
                # 但这需要修改 FastMCP 的底层实现
                #
                # 关键发现：我们可以修改底层服务器的代码，让它能够处理生成器
                # 但我们需要找到底层服务器的代码位置
                #
                # 或者，我们可以通过其他方式实现流式传输：
                # 1. 使用 MCP 的进度消息机制（如果存在）
                # 2. 或者直接通过 SSE 连接发送消息（需要访问底层传输层）
                
                # 实现真正的流式传输：在新线程中运行事件循环来收集消息
                def streaming_generator():
                    """同步生成器，逐个 yield TextContent，实现真正的流式传输"""
                    import asyncio
                    from mcp.types import TextContent
                    import queue
                    import threading
                    
                    # 创建一个队列来传递值
                    q = queue.Queue()
                    done = False
                    error = None
                    message_count = 0
                    
                    async def collect():
                        nonlocal done, error, message_count
                        try:
                            logger.info(f"[流式传输] 开始收集异步生成器的消息 (工具: {tool_name})")
                            logger.info(f"[流式传输] results 类型: {type(results)}, 是否为异步生成器: {inspect.isasyncgen(results)}")
                            
                            # 检查 results 是否是异步生成器
                            if not inspect.isasyncgen(results):
                                logger.error(f"[流式传输] results 不是异步生成器！类型: {type(results)}")
                                error = ValueError(f"results 不是异步生成器，而是 {type(results)}")
                                return
                            
                            logger.info(f"[流式传输] 开始迭代异步生成器 (工具: {tool_name})")
                            async for item in results:
                                logger.debug(f"[流式传输] 收到 item (工具: {tool_name}), 类型: {type(item)}")
                                
                                # 转换每个 item 为 TextContent
                                if isinstance(item, dict):
                                    text = json.dumps(item, ensure_ascii=False, indent=2)
                                    content = TextContent(type="text", text=text)
                                    q.put(content)
                                    # 记录推送的消息
                                    message_count += 1
                                    logger.info(f"[流式传输] 推送消息 #{message_count} (工具: {tool_name}): {text[:200]}..." if len(text) > 200 else f"[流式传输] 推送消息 #{message_count} (工具: {tool_name}): {text}")
                                elif isinstance(item, TextContent):
                                    q.put(item)
                                    # 记录推送的消息
                                    message_count += 1
                                    text_preview = item.text[:200] if len(item.text) > 200 else item.text
                                    logger.info(f"[流式传输] 推送消息 #{message_count} (工具: {tool_name}): {text_preview}...")
                                else:
                                    content = TextContent(type="text", text=str(item))
                                    q.put(content)
                                    # 记录推送的消息
                                    message_count += 1
                                    text_preview = str(item)[:200] if len(str(item)) > 200 else str(item)
                                    logger.info(f"[流式传输] 推送消息 #{message_count} (工具: {tool_name}): {text_preview}...")
                            
                            logger.info(f"[流式传输] 异步生成器迭代完成 (工具: {tool_name}), 共收集 {message_count} 条消息")
                        except Exception as e:
                            error = e
                            logger.error(f"[流式传输] 收集消息时出错 (工具: {tool_name}): {str(e)}", exc_info=True)
                        finally:
                            done = True
                            q.put(None)  # 结束标记
                            logger.info(f"[流式传输] 消息收集完成 (工具: {tool_name}), 共推送 {message_count} 条消息")
                    
                    # 在新线程中运行事件循环来执行 collect()
                    def run_collect():
                        # 创建新的事件循环
                        new_loop = asyncio.new_event_loop()
                        asyncio.set_event_loop(new_loop)
                        try:
                            logger.info(f"[流式传输] 在新线程中启动事件循环 (工具: {tool_name})")
                            new_loop.run_until_complete(collect())
                        finally:
                            new_loop.close()
                    
                    # 启动收集线程
                    collect_thread = threading.Thread(target=run_collect, daemon=True)
                    collect_thread.start()
                    logger.info(f"[流式传输] 开始流式传输 (工具: {tool_name}), 已启动收集线程")
                    
                    # 从队列中获取值并 yield
                    # 这样底层服务器在迭代时能够实时获取值
                    yielded_count = 0
                    logger.info(f"[流式传输] 开始从队列获取消息 (工具: {tool_name})")
                    while True:
                        try:
                            logger.debug(f"[流式传输] 等待队列中的消息 (工具: {tool_name})...")
                            item = q.get(timeout=0.1)  # 使用超时，避免永久阻塞
                            logger.debug(f"[流式传输] 从队列获取到消息 (工具: {tool_name}), 类型: {type(item)}")
                        except queue.Empty:
                            # 检查线程是否还在运行
                            if done:
                                logger.info(f"[流式传输] 收集完成，没有更多消息 (工具: {tool_name})")
                                break
                            if error:
                                logger.error(f"[流式传输] 收集出错 (工具: {tool_name}): {str(error)}", exc_info=True)
                                raise error
                            # 线程还在运行，继续等待
                            logger.debug(f"[流式传输] 队列为空，但收集线程仍在运行 (工具: {tool_name})")
                            continue
                        
                        if item is None:  # 结束标记
                            logger.info(f"[流式传输] 收到结束标记 (工具: {tool_name})")
                            logger.info(f"[流式传输] 生成器迭代完成 (工具: {tool_name}), 共 yield {yielded_count} 条消息")
                            break
                        if error:
                            logger.error(f"[流式传输] 流式传输出错 (工具: {tool_name}): {str(error)}", exc_info=True)
                            raise error
                        yielded_count += 1
                        logger.info(f"[流式传输] yield 消息 #{yielded_count} (工具: {tool_name})")
                        yield item
                    
                    # 等待收集线程完成
                    collect_thread.join(timeout=5.0)
                    if collect_thread.is_alive():
                        logger.warning(f"[流式传输] 收集线程未在5秒内完成 (工具: {tool_name})")
                    else:
                        logger.info(f"[流式传输] 收集线程已完成 (工具: {tool_name})")
                
                # 返回生成器，让底层服务器迭代它
                # 关键：底层服务器的代码会调用 list(unstructured_content)
                # 但我们的生成器会在被迭代时实时获取值，所以应该能实现流式传输
                return ServerResult(
                    CallToolResult(
                        content=streaming_generator(),  # 传递生成器，让底层服务器迭代它
                        structuredContent=None,
                        isError=False,
                    )
                )
            
            # 对于非异步生成器，使用原始处理逻辑
            unstructured_content: list
            maybe_structured_content: dict | None
            if isinstance(results, tuple) and len(results) == 2:
                unstructured_content, maybe_structured_content = results
            elif isinstance(results, dict):
                maybe_structured_content = results
                unstructured_content = [TextContent(type="text", text=json.dumps(results, indent=2))]
            elif hasattr(results, "__iter__") and not inspect.isasyncgen(results):
                unstructured_content = list(results)
                maybe_structured_content = None
            else:
                return mcp._mcp_server._make_error_result(f"Unexpected return type from tool: {type(results).__name__}")
            
            # output validation
            if tool and tool.outputSchema is not None:
                if maybe_structured_content is None:
                    return mcp._mcp_server._make_error_result(
                        "Output validation error: outputSchema defined but no structured output returned"
                    )
                else:
                    try:
                        jsonschema.validate(instance=maybe_structured_content, schema=tool.outputSchema)
                    except jsonschema.ValidationError as e:
                        return mcp._mcp_server._make_error_result(f"Output validation error: {e.message}")
            
            # result
            # 关键：底层服务器的代码会调用 list(unstructured_content)
            # 如果 unstructured_content 是生成器，list() 会立即消费它
            # 但我们的生成器会在迭代时实时获取值，所以应该能实现流式传输
            # 
            # 实现真正的流式传输：不调用 list()，直接传递生成器
            # 但底层服务器的代码会调用 list()，所以我们需要一个特殊的生成器
            # 它在被迭代时能够实时发送 SSE 消息
            #
            # 关键：我们需要修改底层服务器的代码，让它不调用 list()
            # 而是直接传递生成器给 SSE 传输层
            # 但这需要修改系统库文件，不应该直接修改
            #
            # 最终方案：返回生成器，让底层服务器迭代它
            # 虽然底层服务器会调用 list()，但我们的生成器会在迭代时实时获取值
            # 这样至少能保证值是在迭代时实时获取的，而不是预先收集的
            if hasattr(unstructured_content, "__iter__") and not isinstance(unstructured_content, (list, tuple, str, bytes)):
                # 如果是生成器，直接传递，不转换为列表
                # 底层服务器的代码会调用 list()，但我们的生成器会在迭代时实时获取值
                return ServerResult(
                    CallToolResult(
                        content=unstructured_content,  # 直接传递生成器，不转换为列表
                        structuredContent=maybe_structured_content,
                        isError=False,
                    )
                )
            else:
                # 如果不是生成器，转换为列表
                return ServerResult(
                    CallToolResult(
                        content=list(unstructured_content),
                        structuredContent=maybe_structured_content,
                        isError=False,
                    )
                )
        except Exception as e:
            handler_elapsed = time.time() - handler_start_time
            logger.error(f"[流式传输] ❌ Handler 异常 (工具: {tool_name if 'tool_name' in locals() else 'unknown'}, 耗时: {handler_elapsed:.3f}秒): {str(e)}", exc_info=True)
            return mcp._mcp_server._make_error_result(str(e))
    
    # 替换 handler
    mcp._mcp_server.request_handlers[CallToolRequest] = _streaming_handler

# 如果 _setup_handlers 已经被调用，我们需要重新注册 handler
# 但通常 _setup_handlers 在 run() 时才被调用，所以这里应该没问题


def download_to_temp_file(url: str) -> str:
    """
    下载 HTTP/HTTPS URL 到临时文件（以二进制模式下载，确保编码正确）
    
    Args:
        url: HTTP/HTTPS URL
        
    Returns:
        临时文件路径
        
    Raises:
        Exception: 如果下载失败
    """
    logger.info(f"Starting file download: {url}")
    start_time = time.time()
    
    try:
        # 从 URL 获取文件扩展名
        parsed_url = urlparse(url)
        url_path = parsed_url.path
        # 获取文件扩展名，如果没有则使用 .fasta 作为默认扩展名
        ext = os.path.splitext(url_path)[1] or '.fasta'
        
        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
        temp_file_path = temp_file.name
        temp_file.close()
        
        # 使用 requests 库（如果可用）来更好地处理编码和响应头
        if HAS_REQUESTS:
            response = requests.get(url, stream=True, timeout=30)
            response.raise_for_status()
            
            # 记录响应头信息（用于调试）
            content_type = response.headers.get('Content-Type', '')
            content_encoding = response.headers.get('Content-Encoding', '')
            logger.debug(f"Response headers - Content-Type: {content_type}, Content-Encoding: {content_encoding}")
            
            # 以二进制模式写入文件（确保不进行任何编码转换）
            with open(temp_file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
        else:
            # 回退到 urllib（也以二进制模式下载）
            urllib.request.urlretrieve(url, temp_file_path)
        
        download_time = time.time() - start_time
        file_size = os.path.getsize(temp_file_path) / (1024 * 1024)  # MB
        logger.info(f"File download completed: {temp_file_path} ({file_size:.2f} MB, elapsed: {download_time:.2f}s)")
        
        return temp_file_path
    except Exception as e:
        logger.error(f"File download failed: {url} - {str(e)}")
        raise Exception(f"Failed to download URL {url}: {str(e)}")


def _check_excel_dependencies() -> Tuple[bool, str]:
    """
    检查读取 Excel 文件所需的依赖是否已安装
    
    Returns:
        (is_available, error_message) - 依赖是否可用和错误消息（如果不可用）
    """
    try:
        import openpyxl
        return True, ""
    except ImportError:
        error_msg = (
            "Missing required dependency 'openpyxl' for reading Excel files. "
            "Please install it using one of the following commands:\n"
            "  - pip install openpyxl\n"
            "  - conda install openpyxl"
        )
        return False, error_msg


def _detect_file_type(file_path: Path) -> str:
    """
    通过文件内容（magic bytes）检测文件的实际类型
    
    Args:
        file_path: 文件路径
        
    Returns:
        文件扩展名（带点号），如 '.csv', '.xlsx', '.xls'
    """
    try:
        with open(file_path, 'rb') as f:
            # 读取文件开头几个字节
            header = f.read(8)
        
        # Excel 文件 (XLSX) 是 ZIP 格式，以 PK 开头
        if header.startswith(b'PK\x03\x04'):
            # 进一步检查是否为 Excel
            try:
                import zipfile
                with zipfile.ZipFile(file_path, 'r') as zip_file:
                    file_list = zip_file.namelist()
                    # 检查是否包含典型的 Excel 文件结构
                    if any(name.startswith('xl/') or name.startswith('[Content_Types].xml') for name in file_list):
                        logger.debug("Detected Excel file (XLSX) by magic bytes and structure")
                        return '.xlsx'
                    # 如果是 ZIP 格式但不是 Excel，可能是其他 Office 文档，但先返回 xlsx 让读取尝试
                    logger.debug("Detected ZIP-based file, assuming Excel format")
                    return '.xlsx'
            except Exception as e:
                logger.debug(f"ZIP file check failed: {str(e)}, may not be Excel")
                # 如果无法打开 ZIP，可能是损坏的文件，但仍然可能是 Excel
                return '.xlsx'
        
        # Excel 97-2003 (XLS) 格式
        if header[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
            logger.debug("Detected Excel file (XLS) by magic bytes")
            return '.xls'
        
        # CSV 通常是纯文本，检查是否可读
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                first_line = f.readline()
                # CSV 通常包含逗号、分号或制表符作为分隔符
                if any(sep in first_line for sep in [',', ';', '\t']):
                    logger.debug("Detected CSV file by content analysis")
                    return '.csv'
        except:
            pass
        
        # 默认返回原始扩展名
        return file_path.suffix.lower() or '.csv'
        
    except Exception as e:
        logger.warning(f"File type detection failed: {str(e)}, using extension")
        return file_path.suffix.lower() or '.csv'


def _read_csv_robust(csv_path: Path) -> pd.DataFrame:
    """
    使用多种策略尝试读取 CSV 文件，自动修正常见格式问题
    
    Args:
        csv_path: CSV 文件路径
        
    Returns:
        读取成功的数据框
        
    Raises:
        ValueError: 如果所有读取策略都失败
    """
    import csv
    
    # 检查 pandas 版本，确定是否支持 on_bad_lines 参数
    pandas_version = pd.__version__
    has_on_bad_lines = tuple(map(int, pandas_version.split('.')[:2])) >= (1, 3)
    
    # 策略列表：按优先级尝试
    strategies = []
    
    # 策略1: 标准读取，自动检测分隔符（兼容新旧版本）
    for encoding in ['utf-8', 'gbk', 'latin-1']:
        base_strategy = {'encoding': encoding, 'sep': None, 'engine': 'python'}
        if has_on_bad_lines:
            base_strategy['on_bad_lines'] = 'skip'
        else:
            base_strategy['error_bad_lines'] = False
            base_strategy['warn_bad_lines'] = False
        strategies.append(base_strategy)
    
    # 策略2: 指定分隔符，容错处理
    for sep in [',', '\t', ';']:
        for encoding in ['utf-8', 'gbk']:
            strategy = {'encoding': encoding, 'sep': sep}
            if has_on_bad_lines:
                strategy['on_bad_lines'] = 'skip'
            else:
                strategy['error_bad_lines'] = False
                strategy['warn_bad_lines'] = False
            strategies.append(strategy)
    
    # 策略3: 宽松的引号处理
    for encoding in ['utf-8', 'gbk']:
        strategy = {'encoding': encoding, 'sep': ',', 'quoting': csv.QUOTE_MINIMAL}
        if has_on_bad_lines:
            strategy['on_bad_lines'] = 'skip'
        else:
            strategy['error_bad_lines'] = False
            strategy['warn_bad_lines'] = False
        strategies.append(strategy)
    
    # 策略4: 忽略引号，按分隔符解析
    for encoding in ['utf-8', 'gbk']:
        strategy = {'encoding': encoding, 'sep': ',', 'quoting': csv.QUOTE_NONE}
        if has_on_bad_lines:
            strategy['on_bad_lines'] = 'skip'
        else:
            strategy['error_bad_lines'] = False
            strategy['warn_bad_lines'] = False
        strategies.append(strategy)
    
    # 策略5: 使用 python 引擎，宽松处理
    for encoding in ['utf-8', 'gbk', 'latin-1']:
        strategy = {'encoding': encoding, 'sep': ',', 'engine': 'python'}
        if has_on_bad_lines:
            strategy['on_bad_lines'] = 'skip'
        else:
            strategy['error_bad_lines'] = False
            strategy['warn_bad_lines'] = False
        strategies.append(strategy)
    
    last_error = None
    successful_strategy = None
    
    for i, strategy in enumerate(strategies, 1):
        try:
            logger.debug(f"Trying CSV reading strategy {i}/{len(strategies)}: encoding={strategy.get('encoding')}, sep={strategy.get('sep')}")
            df = pd.read_csv(csv_path, **strategy)
            if not df.empty and len(df.columns) > 0:
                successful_strategy = i
                logger.info(f"Successfully read CSV using strategy {i}/{len(strategies)} "
                          f"(encoding: {strategy.get('encoding', 'default')}, "
                          f"sep: {repr(strategy.get('sep', 'auto'))})")
                return df
        except TypeError as e:
            # 参数不支持，跳过
            logger.debug(f"Strategy {i} skipped: unsupported parameter - {str(e)}")
            continue
        except UnicodeDecodeError as e:
            last_error = e
            logger.debug(f"Strategy {i} failed: encoding error - {str(e)}")
            continue
        except Exception as e:
            last_error = e
            error_msg = str(e)
            # 过滤掉一些常见的临时错误信息，避免日志过多
            if any(keyword in error_msg.lower() for keyword in ['expected', 'quote', 'delimiter', 'parse']):
                logger.debug(f"Strategy {i} failed: parsing error")
            else:
                logger.debug(f"Strategy {i} failed: {error_msg}")
            continue
    
    # 如果所有策略都失败，尝试手动逐行解析（最后的手段）
    logger.warning(f"All {len(strategies)} standard CSV reading strategies failed, attempting manual parsing...")
    if last_error:
        logger.debug(f"Last error: {str(last_error)}")
    return _read_csv_manual(csv_path)


def _read_csv_manual(csv_path: Path) -> pd.DataFrame:
    """
    手动逐行解析 CSV 文件，处理格式不规范的情况
    
    Args:
        csv_path: CSV 文件路径
        
    Returns:
        解析后的数据框
    """
    import csv
    
    logger.info("Starting manual CSV parsing with error recovery...")
    
    # 尝试不同的编码读取文件内容
    content = None
    for encoding in ['utf-8', 'gbk', 'latin-1']:
        try:
            with open(csv_path, 'r', encoding=encoding) as f:
                content = f.read()
            logger.debug(f"Successfully read file content with encoding: {encoding}")
            break
        except UnicodeDecodeError:
            continue
    
    if content is None:
        raise ValueError("Unable to read file with any encoding")
    
    # 清理内容：移除 BOM、修复常见的引号问题
    content = content.lstrip('\ufeff')  # 移除 UTF-8 BOM
    
    # 尝试使用 csv.Sniffer 检测分隔符
    try:
        sample = content[:1024] if len(content) > 1024 else content
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t')
        delimiter = dialect.delimiter
        logger.info(f"Detected delimiter: '{delimiter}'")
    except:
        delimiter = ','
        logger.warning(f"Could not detect delimiter, using default: ','")
    
    # 手动解析行，尝试多种解析方式
    lines = content.split('\n')
    rows = []
    headers = None
    skip_count = 0
    
    # 尝试不同的解析配置
    parse_configs = [
        {'delimiter': delimiter, 'quoting': csv.QUOTE_MINIMAL},
        {'delimiter': delimiter, 'quoting': csv.QUOTE_NONE},
        {'delimiter': delimiter, 'quoting': csv.QUOTE_MINIMAL, 'escapechar': '\\'},
        {'delimiter': delimiter, 'quoting': csv.QUOTE_ALL},
    ]
    
    for line_idx, line in enumerate(lines):
        line = line.strip()
        if not line or line.startswith('#'):  # 跳过空行和注释
            continue
        
        row = None
        last_parse_error = None
        
        # 尝试不同的解析配置
        for config in parse_configs:
            try:
                reader = csv.reader([line], **config)
                row = next(reader)
                if row and len(row) > 0:
                    break
            except Exception as e:
                last_parse_error = e
                continue
        
        # 如果所有配置都失败，尝试简单的分隔符分割（最后手段）
        if row is None or len(row) == 0:
            try:
                # 简单的分隔符分割，清理引号
                row = [col.strip().strip('"').strip("'") for col in line.split(delimiter)]
            except:
                logger.debug(f"Failed to parse line {line_idx + 1} with all methods: {str(last_parse_error)}, skipping")
                skip_count += 1
                continue
        
        if not row or len(row) == 0:
            skip_count += 1
            continue
        
        if headers is None:
            headers = row
            logger.info(f"Detected headers ({len(headers)} columns): {headers}")
        else:
            # 处理列数不一致的情况
            if len(row) < len(headers):
                # 列数不足，补齐空字符串
                row = row + [''] * (len(headers) - len(row))
            elif len(row) > len(headers):
                # 列数过多，只取前几列
                row = row[:len(headers)]
            rows.append(row)
    
    if headers is None or not rows:
        raise ValueError(f"Could not parse CSV file: no valid headers or rows found")
    
    if skip_count > 0:
        logger.warning(f"Skipped {skip_count} malformed rows during manual parsing")
    
    logger.info(f"Manually parsed {len(rows)} valid rows from {len(headers)} columns")
    
    # 确保所有行的列数一致，补齐缺失列或截断多余列
    max_cols = max(len(row) for row in rows) if rows else len(headers)
    num_cols = min(max_cols, len(headers))
    
    # 统一行的长度
    normalized_rows = []
    for row in rows:
        if len(row) < num_cols:
            # 补齐缺失的列
            row = row + [''] * (num_cols - len(row))
        elif len(row) > num_cols:
            # 截断多余的列
            row = row[:num_cols]
        normalized_rows.append(row)
    
    # 创建数据框
    df = pd.DataFrame(normalized_rows, columns=headers[:num_cols])
    logger.debug(f"Created DataFrame with {len(df)} rows and {len(df.columns)} columns")
    return df


def _validate_csv_file(file_path: Path) -> Tuple[bool, str]:
    """
    验证 CSV 文件是否符合转换要求
    
    允许空文件（文件大小为0），空文件被认为是有效的。
    
    Args:
        file_path: CSV 文件路径
        
    Returns:
        (is_valid, error_message) - 文件是否有效和错误消息（如果无效）
    """
    try:
        # 检查文件是否存在
        if not file_path.exists():
            return False, f"File does not exist: {file_path}"
        
        # 检查文件大小（允许空文件）
        file_size = file_path.stat().st_size
        if file_size == 0:
            logger.debug(f"CSV file is empty: {file_path} (allowed)")
            # 空文件仍然被认为是有效的（允许后续处理）
            return True, ""
        
        # 检查文件扩展名
        if file_path.suffix.lower() != '.csv':
            return False, f"File extension is not .csv: {file_path.suffix}"
        
        # 检查文件内容（尝试读取第一行）
        # 尝试多种编码，与 _read_csv_robust 保持一致
        encodings = ['utf-8', 'gbk', 'latin-1', 'cp1252', 'iso-8859-1']
        file_readable = False
        last_error = None
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    first_line = f.readline()
                    if first_line and first_line.strip():
                        file_readable = True
                        break
            except UnicodeDecodeError as e:
                last_error = e
                continue
            except Exception as e:
                last_error = e
                continue
        
        if not file_readable:
            return False, f"Cannot read file with common encodings ({', '.join(encodings)}): {str(last_error) if last_error else 'Unknown error'}"
        
        # 注意：Excel 文件的检测和处理现在在 convert_csv_to_fasta 函数中进行，
        # 该函数会自动检测文件类型并使用相应的转换器
        
        return True, ""
    except Exception as e:
        return False, f"Error validating CSV file: {str(e)}"


def _validate_excel_file(file_path: Path) -> Tuple[bool, str]:
    """
    验证 Excel 文件是否符合转换要求
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        (is_valid, error_message) - 文件是否有效和错误消息（如果无效）
    """
    try:
        # 检查文件是否存在
        if not file_path.exists():
            return False, f"File does not exist: {file_path}"
        
        # 检查文件大小
        file_size = file_path.stat().st_size
        if file_size == 0:
            return False, "File is empty"
        
        # 检查 Excel 依赖
        excel_available, excel_error_msg = _check_excel_dependencies()
        if not excel_available:
            return False, excel_error_msg
        
        # 检查文件类型（通过 magic bytes）- 这是主要检查，优先于扩展名检查
        actual_type = _detect_file_type(file_path)
        if actual_type not in ['.xlsx', '.xls']:
            return False, f"File is actually {actual_type} format, not Excel. Please use the appropriate converter."
        
        # 检查文件扩展名（次要检查，如果实际类型是 Excel 但扩展名不匹配，给出警告但允许通过）
        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            logger.warning(f"File extension ({file_path.suffix}) does not match actual file type ({actual_type}), but proceeding with conversion based on actual file type")
        
        # 尝试打开文件验证
        try:
            df = pd.read_excel(file_path, nrows=1)  # 只读取第一行来验证
            if df.empty:
                return False, "Excel file appears to be empty"
        except Exception as e:
            return False, f"Cannot read Excel file: {str(e)}"
        
        return True, ""
    except Exception as e:
        return False, f"Error validating Excel file: {str(e)}"


def _convert_csv_to_fasta(input_file: Union[str, Path], output_file: Union[str, Path]) -> None:
    """
    将 CSV 文件转换为 FASTA 格式
    
    自动检测列名，支持多种常见的列名变体。
    对格式不规范的 CSV 文件具有强大的容错能力。
    
    Args:
        input_file: 输入 CSV 文件路径
        output_file: 输出的 FASTA 文件路径
        
    Raises:
        ValueError: 如果文件验证失败、找不到必需的列或数据格式错误
        Exception: 如果文件读取或写入失败
    """
    input_path = Path(input_file)
    output_path = Path(output_file)
    
    logger.info(f"Starting CSV to FASTA conversion: {input_path} -> {output_path}")
    start_time = time.time()
    
    # 验证 CSV 文件
    is_valid, error_msg = _validate_csv_file(input_path)
    if not is_valid:
        logger.error(f"CSV file validation failed: {error_msg}")
        raise ValueError(f"CSV file validation failed: {error_msg}")
    
    logger.info("CSV file validation passed")
    
    try:
        # 读取 CSV 文件
        df = _read_csv_robust(input_path)
        
        # 检查数据框是否为空
        if df.empty:
            error_msg = "Input CSV file is empty or contains no data"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.info(f"Loaded {len(df)} rows, {len(df.columns)} columns from CSV file")
        logger.debug(f"Column names: {list(df.columns)}")
        
        # 自动检测 ID 列和序列列
        id_column = None
        seq_column = None
        
        # ID 列的常见名称变体（按优先级排序，不区分大小写）
        # 优先使用 main_name（抗体数据常用）
        id_candidates = ['main_name', 'id', 'sequence_id', 'seq_id', 'name', 'identifier', 'sample_id', 'seq_name']
        # 序列列的常见名称变体（按优先级排序，不区分大小写）
        # 优先使用 Heavy_DNA（抗体重链DNA序列），避免使用 variant_seq（抗原序列）
        seq_candidates = ['heavy_dna', 'heavy_dna_sequence', 'sequence', 'seq', 'nucleotide', 'nuc', 'cdr3', 'junction', 'dna_sequence']
        # 明确排除的序列列（这些是抗原序列，不是抗体序列）
        excluded_seq_columns = ['variant_seq', 'antigen_seq', 'variant', 'antigen_sequence']
        
        # 查找 ID 列（精确匹配，不区分大小写，按优先级）
        for candidate in id_candidates:
            for col in df.columns:
                col_lower = col.lower().strip()
                if col_lower == candidate.lower():
                    id_column = col
                    logger.info(f"Auto-detected ID column: '{col}'")
                    break
            if id_column:
                break
        
        # 如果没有找到，尝试包含关键词的列名（但排除已排除的列）
        if id_column is None:
            for col in df.columns:
                col_lower = col.lower().strip()
                if any(keyword in col_lower for keyword in ['id', 'name', 'identifier']):
                    id_column = col
                    logger.info(f"Auto-detected ID column: '{col}'")
                    break
        
        # 查找序列列（精确匹配，不区分大小写，按优先级，排除抗原序列列）
        for candidate in seq_candidates:
            for col in df.columns:
                col_lower = col.lower().strip()
                # 跳过排除的列
                if any(excluded in col_lower for excluded in excluded_seq_columns):
                    continue
                if col_lower == candidate.lower():
                    seq_column = col
                    logger.info(f"Auto-detected sequence column: '{col}'")
                    break
            if seq_column:
                break
        
        # 如果没有找到，尝试包含关键词的列名（但排除抗原序列列）
        if seq_column is None:
            for col in df.columns:
                col_lower = col.lower().strip()
                # 明确排除抗原序列列
                if any(excluded in col_lower for excluded in excluded_seq_columns):
                    logger.debug(f"Skipping excluded sequence column: '{col}' (antigen sequence, not antibody sequence)")
                    continue
                if any(keyword in col_lower for keyword in ['sequence', 'seq', 'nucleotide', 'dna']):
                    seq_column = col
                    logger.info(f"Auto-detected sequence column: '{col}'")
                    break
        
        # 验证必需的列是否存在
        if id_column is None:
            error_msg = f"Could not identify ID column. Available columns: {list(df.columns)}. " \
                       f"Expected column names: {id_candidates}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if seq_column is None:
            error_msg = f"Could not identify sequence column. Available columns: {list(df.columns)}. " \
                       f"Expected column names: {seq_candidates}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # 检查并清理数据
        df = df[[id_column, seq_column]].copy()
        df = df.dropna(subset=[id_column, seq_column])  # 移除空值行
        
        # 转换序列为字符串并清理
        df[seq_column] = df[seq_column].astype(str).str.strip()
        df[id_column] = df[id_column].astype(str).str.strip()
        
        # 过滤掉空序列
        df = df[df[seq_column].str.len() > 0]
        df = df[df[id_column].str.len() > 0]
        
        if df.empty:
            error_msg = "No valid sequences found after filtering"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.info(f"Valid sequences after filtering: {len(df)}")
        
        # 写入 FASTA 文件
        logger.debug(f"Writing FASTA file: {output_path}")
        try:
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, 'w') as f:
                for _, row in df.iterrows():
                    seq_id = str(row[id_column]).strip()
                    sequence = str(row[seq_column]).strip()
                    
                    # 确保序列 ID 以 '>' 开头
                    if not seq_id.startswith('>'):
                        seq_id = f">{seq_id}"
                    
                    # 写入 FASTA 格式
                    f.write(f"{seq_id}\n{sequence}\n")
            
            # 验证文件是否成功写入
            if not output_path.exists():
                error_msg = f"FASTA file was not created: {output_path}"
                logger.error(error_msg)
                raise IOError(error_msg)
            
            if output_path.stat().st_size == 0:
                error_msg = f"FASTA file is empty: {output_path}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
        except Exception as e:
            error_msg = f"Failed to write FASTA file to {output_path}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise IOError(error_msg) from e
        
        conversion_time = time.time() - start_time
        file_size = output_path.stat().st_size / 1024  # KB
        logger.info(f"CSV to FASTA conversion completed: {len(df)} sequences, "
                   f"output size: {file_size:.2f} KB, elapsed: {conversion_time:.2f}s")
        
    except ValueError as e:
        logger.error(f"Value error during CSV to FASTA conversion: {str(e)}", exc_info=True)
        raise
    except Exception as e:
        error_msg = f"Unexpected error during CSV to FASTA conversion: {str(e)}"
        logger.error(error_msg, exc_info=True)
        raise Exception(error_msg) from e


def _convert_excel_to_fasta(input_file: Union[str, Path], output_file: Union[str, Path], expected_format: Optional[str] = None) -> None:
    """
    将 Excel 文件转换为 FASTA 格式
    
    自动检测列名，支持多种常见的列名变体。
    支持 XLSX 和 XLS 格式。
    
    Args:
        input_file: 输入 Excel 文件路径（XLSX 或 XLS）
        output_file: 输出的 FASTA 文件路径
        expected_format: 期望的文件格式（'.xlsx' 或 '.xls'），如果提供则进行格式验证
        
    Raises:
        ValueError: 如果文件验证失败、找不到必需的列或数据格式错误
        ImportError: 如果缺少 Excel 读取依赖
        Exception: 如果文件读取或写入失败
    """
    input_path = Path(input_file)
    output_path = Path(output_file)
    
    logger.info(f"Starting Excel to FASTA conversion: {input_path} -> {output_path}")
    start_time = time.time()
    
    # 验证 Excel 文件
    is_valid, error_msg = _validate_excel_file(input_path)
    if not is_valid:
        logger.error(f"Excel file validation failed: {error_msg}")
        if "dependency" in error_msg.lower() or "openpyxl" in error_msg.lower():
            raise ImportError(f"Excel file validation failed: {error_msg}")
        raise ValueError(f"Excel file validation failed: {error_msg}")
    
    # 如果指定了期望格式，验证文件格式是否匹配
    if expected_format:
        actual_type = _detect_file_type(input_path)
        if actual_type != expected_format:
            error_msg = f"File format mismatch: expected {expected_format}, but file is {actual_type}"
            logger.error(error_msg)
            raise ValueError(error_msg)
    
    logger.info("Excel file validation passed")
    
    try:
        # 读取 Excel 文件
        df = pd.read_excel(input_path)
        
        # 检查数据框是否为空
        if df.empty:
            error_msg = "Input Excel file is empty or contains no data"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.info(f"Loaded {len(df)} rows, {len(df.columns)} columns from Excel file")
        logger.debug(f"Column names: {list(df.columns)}")
        
        # 自动检测 ID 列和序列列
        id_column = None
        seq_column = None
        
        # ID 列的常见名称变体（按优先级排序，不区分大小写）
        # 优先使用 main_name（抗体数据常用）
        id_candidates = ['main_name', 'id', 'sequence_id', 'seq_id', 'name', 'identifier', 'sample_id', 'seq_name']
        # 序列列的常见名称变体（按优先级排序，不区分大小写）
        # 优先使用 Heavy_DNA（抗体重链DNA序列），避免使用 variant_seq（抗原序列）
        seq_candidates = ['heavy_dna', 'heavy_dna_sequence', 'sequence', 'seq', 'nucleotide', 'nuc', 'cdr3', 'junction', 'dna_sequence']
        # 明确排除的序列列（这些是抗原序列，不是抗体序列）
        excluded_seq_columns = ['variant_seq', 'antigen_seq', 'variant', 'antigen_sequence']
        
        # 查找 ID 列（精确匹配，不区分大小写，按优先级）
        for candidate in id_candidates:
            for col in df.columns:
                col_lower = col.lower().strip()
                if col_lower == candidate.lower():
                    id_column = col
                    logger.info(f"Auto-detected ID column: '{col}'")
                    break
            if id_column:
                break
        
        # 如果没有找到，尝试包含关键词的列名（但排除已排除的列）
        if id_column is None:
            for col in df.columns:
                col_lower = col.lower().strip()
                if any(keyword in col_lower for keyword in ['id', 'name', 'identifier']):
                    id_column = col
                    logger.info(f"Auto-detected ID column: '{col}'")
                    break
        
        # 查找序列列（精确匹配，不区分大小写，按优先级，排除抗原序列列）
        for candidate in seq_candidates:
            for col in df.columns:
                col_lower = col.lower().strip()
                # 跳过排除的列
                if any(excluded in col_lower for excluded in excluded_seq_columns):
                    continue
                if col_lower == candidate.lower():
                    seq_column = col
                    logger.info(f"Auto-detected sequence column: '{col}'")
                    break
            if seq_column:
                break
        
        # 如果没有找到，尝试包含关键词的列名（但排除抗原序列列）
        if seq_column is None:
            for col in df.columns:
                col_lower = col.lower().strip()
                # 明确排除抗原序列列
                if any(excluded in col_lower for excluded in excluded_seq_columns):
                    logger.debug(f"Skipping excluded sequence column: '{col}' (antigen sequence, not antibody sequence)")
                    continue
                if any(keyword in col_lower for keyword in ['sequence', 'seq', 'nucleotide', 'dna']):
                    seq_column = col
                    logger.info(f"Auto-detected sequence column: '{col}'")
                    break
        
        # 验证必需的列是否存在
        if id_column is None:
            error_msg = f"Could not identify ID column. Available columns: {list(df.columns)}. " \
                       f"Expected column names: {id_candidates}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        if seq_column is None:
            error_msg = f"Could not identify sequence column. Available columns: {list(df.columns)}. " \
                       f"Expected column names: {seq_candidates}"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # 检查并清理数据
        df = df[[id_column, seq_column]].copy()
        df = df.dropna(subset=[id_column, seq_column])  # 移除空值行
        
        # 转换序列为字符串并清理
        df[seq_column] = df[seq_column].astype(str).str.strip()
        df[id_column] = df[id_column].astype(str).str.strip()
        
        # 过滤掉空序列
        df = df[df[seq_column].str.len() > 0]
        df = df[df[id_column].str.len() > 0]
        
        if df.empty:
            error_msg = "No valid sequences found after filtering"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.info(f"Valid sequences after filtering: {len(df)}")
        
        # 写入 FASTA 文件
        logger.debug(f"Writing FASTA file: {output_path}")
        try:
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            with open(output_path, 'w') as f:
                for _, row in df.iterrows():
                    seq_id = str(row[id_column]).strip()
                    sequence = str(row[seq_column]).strip()
                    
                    # 确保序列 ID 以 '>' 开头
                    if not seq_id.startswith('>'):
                        seq_id = f">{seq_id}"
                    
                    # 写入 FASTA 格式
                    f.write(f"{seq_id}\n{sequence}\n")
            
            # 验证文件是否成功写入
            if not output_path.exists():
                error_msg = f"FASTA file was not created: {output_path}"
                logger.error(error_msg)
                raise IOError(error_msg)
            
            if output_path.stat().st_size == 0:
                error_msg = f"FASTA file is empty: {output_path}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
        except Exception as e:
            error_msg = f"Failed to write FASTA file to {output_path}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise IOError(error_msg) from e
        
        conversion_time = time.time() - start_time
        file_size = output_path.stat().st_size / 1024  # KB
        logger.info(f"Excel to FASTA conversion completed: {len(df)} sequences, "
                   f"output size: {file_size:.2f} KB, elapsed: {conversion_time:.2f}s")
        
    except ImportError as e:
        # 依赖缺失错误，直接抛出并提供清晰的错误消息
        error_str = str(e)
        if 'openpyxl' in error_str.lower():
            excel_available, excel_error_msg = _check_excel_dependencies()
            if not excel_available:
                error_msg = f"Excel file detected but required dependency is missing. {excel_error_msg}"
                logger.error(error_msg)
                raise ImportError(error_msg) from e
        raise
    except ValueError as e:
        logger.error(f"Value error during Excel to FASTA conversion: {str(e)}", exc_info=True)
        raise
    except Exception as e:
        error_str = str(e)
        # 检查是否是依赖缺失的错误
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            excel_available, excel_error_msg = _check_excel_dependencies()
            if not excel_available:
                error_msg = f"Excel file detected but required dependency is missing. {excel_error_msg}"
                logger.error(error_msg)
                raise ImportError(error_msg) from e
        error_msg = f"Unexpected error during Excel to FASTA conversion: {str(e)}"
        logger.error(error_msg, exc_info=True)
        raise Exception(error_msg) from e


def _convert_excel_to_csv(input_file: Union[str, Path], output_file: Union[str, Path], expected_format: Optional[str] = None) -> None:
    """
    将 Excel 文件转换为 CSV 格式
    
    支持 XLSX 和 XLS 格式。
    保留所有列和数据，不进行任何过滤。
    
    Args:
        input_file: 输入 Excel 文件路径（XLSX 或 XLS）
        output_file: 输出的 CSV 文件路径
        expected_format: 期望的文件格式（'.xlsx' 或 '.xls'），如果提供则进行格式验证
        
    Raises:
        ValueError: 如果文件验证失败或数据格式错误
        ImportError: 如果缺少 Excel 读取依赖
        Exception: 如果文件读取或写入失败
    """
    input_path = Path(input_file)
    output_path = Path(output_file)
    
    logger.info(f"Starting Excel to CSV conversion: {input_path} -> {output_path}")
    start_time = time.time()
    
    # 验证 Excel 文件
    is_valid, error_msg = _validate_excel_file(input_path)
    if not is_valid:
        logger.error(f"Excel file validation failed: {error_msg}")
        if "dependency" in error_msg.lower() or "openpyxl" in error_msg.lower():
            raise ImportError(f"Excel file validation failed: {error_msg}")
        raise ValueError(f"Excel file validation failed: {error_msg}")
    
    # 如果指定了期望格式，验证文件格式是否匹配
    if expected_format:
        actual_type = _detect_file_type(input_path)
        if actual_type != expected_format:
            error_msg = f"File format mismatch: expected {expected_format}, but file is {actual_type}"
            logger.error(error_msg)
            raise ValueError(error_msg)
    
    logger.info("Excel file validation passed")
    
    try:
        # 读取 Excel 文件
        df = pd.read_excel(input_path)
        
        # 检查数据框是否为空
        if df.empty:
            error_msg = "Input Excel file is empty or contains no data"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.info(f"Loaded {len(df)} rows, {len(df.columns)} columns from Excel file")
        logger.debug(f"Column names: {list(df.columns)}")
        
        # 写入 CSV 文件
        logger.debug(f"Writing CSV file: {output_path}")
        try:
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 保存为 CSV，使用 UTF-8 编码
            df.to_csv(output_path, index=False, encoding='utf-8')
            
            # 验证文件是否成功写入
            if not output_path.exists():
                error_msg = f"CSV file was not created: {output_path}"
                logger.error(error_msg)
                raise IOError(error_msg)
            
            if output_path.stat().st_size == 0:
                error_msg = f"CSV file is empty: {output_path}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
        except Exception as e:
            error_msg = f"Failed to write CSV file to {output_path}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise IOError(error_msg) from e
        
        conversion_time = time.time() - start_time
        file_size = output_path.stat().st_size / 1024  # KB
        logger.info(f"Excel to CSV conversion completed: {len(df)} rows, {len(df.columns)} columns, "
                   f"output size: {file_size:.2f} KB, elapsed: {conversion_time:.2f}s")
        
    except ImportError as e:
        # 依赖缺失错误，直接抛出并提供清晰的错误消息
        error_str = str(e)
        if 'openpyxl' in error_str.lower():
            excel_available, excel_error_msg = _check_excel_dependencies()
            if not excel_available:
                error_msg = f"Excel file detected but required dependency is missing. {excel_error_msg}"
                logger.error(error_msg)
                raise ImportError(error_msg) from e
        raise
    except ValueError as e:
        logger.error(f"Value error during Excel to CSV conversion: {str(e)}", exc_info=True)
        raise
    except Exception as e:
        error_str = str(e)
        # 检查是否是依赖缺失的错误
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            excel_available, excel_error_msg = _check_excel_dependencies()
            if not excel_available:
                error_msg = f"Excel file detected but required dependency is missing. {excel_error_msg}"
                logger.error(error_msg)
                raise ImportError(error_msg) from e
        error_msg = f"Unexpected error during Excel to CSV conversion: {str(e)}"
        logger.error(error_msg, exc_info=True)
        raise Exception(error_msg) from e


def _convert_csv_to_excel(input_file: Union[str, Path], output_file: Union[str, Path], output_format: str = 'xlsx') -> None:
    """
    将 CSV 文件转换为 Excel 格式
    
    支持输出为 XLSX 或 XLS 格式。
    保留所有列和数据，不进行任何过滤。
    
    Args:
        input_file: 输入 CSV 文件路径
        output_file: 输出的 Excel 文件路径
        output_format: 输出格式，'xlsx' 或 'xls'（默认: 'xlsx'）
        
    Raises:
        ValueError: 如果文件验证失败或数据格式错误
        ImportError: 如果缺少 Excel 写入依赖
        Exception: 如果文件读取或写入失败
    """
    input_path = Path(input_file)
    output_path = Path(output_file)
    
    logger.info(f"Starting CSV to Excel ({output_format.upper()}) conversion: {input_path} -> {output_path}")
    start_time = time.time()
    
    # 验证 CSV 文件
    is_valid, error_msg = _validate_csv_file(input_path)
    if not is_valid:
        logger.error(f"CSV file validation failed: {error_msg}")
        raise ValueError(f"CSV file validation failed: {error_msg}")
    
    logger.info("CSV file validation passed")
    
    # 验证输出格式
    if output_format.lower() not in ['xlsx', 'xls']:
        raise ValueError(f"Invalid output format: {output_format}. Must be 'xlsx' or 'xls'")
    
    # 检查 Excel 依赖
    excel_available, excel_error_msg = _check_excel_dependencies()
    if not excel_available:
        raise ImportError(f"Excel file writing requires dependency. {excel_error_msg}")
    
    try:
        # 读取 CSV 文件
        df = _read_csv_robust(input_path)
        
        # 检查数据框是否为空
        if df.empty:
            error_msg = "Input CSV file is empty or contains no data"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.info(f"Loaded {len(df)} rows, {len(df.columns)} columns from CSV file")
        logger.debug(f"Column names: {list(df.columns)}")
        
        # 清理数据：处理可能导致 Excel 写入问题的特殊字符
        # Excel 不允许的控制字符：\x00-\x08, \x0B, \x0C, \x0E-\x1F
        import re
        
        def clean_excel_string(s):
            """清理字符串中的 Excel 非法字符"""
            if not isinstance(s, str):
                return s
            # 移除所有 Excel 不允许的控制字符（\x00-\x1F 除了 \x09 TAB, \x0A LF, \x0D CR）
            # 但为了安全，我们移除所有控制字符，包括 DEL (\x7F)
            cleaned = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', s)
            # 将 CR 和 LF 替换为空格（避免换行问题）
            cleaned = cleaned.replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
            # 移除多个连续空格
            cleaned = re.sub(r' +', ' ', cleaned).strip()
            # 检查是否包含明显的二进制数据模式（如 ZIP 文件头 "PK\x03\x04"）
            # 或者包含大量不可打印字符
            if len(cleaned) > 0:
                # 检查是否包含 ZIP 文件头模式
                if cleaned.startswith('PK') and len(cleaned) > 2:
                    try:
                        # 检查第三个字符是否是控制字符（ZIP 文件头的特征）
                        if ord(cleaned[2]) < 32 or cleaned[2] in ['\x03', '\x04']:
                            logger.warning(f"Detected binary data pattern (ZIP header) in string, replacing with empty string")
                            return ''
                    except:
                        pass
                # 检查是否包含大量真正的控制字符（只检测 ASCII 控制字符，不包括 Unicode 字符）
                # 注意：Unicode 字符（如中文）的 ord 值 > 126，但它们是合法的可打印字符
                # 我们只检测真正的控制字符：\x00-\x1F 和 \x7F (DEL)
                # 由于前面已经清理过，这里主要检测是否还有残留的控制字符
                control_char_count = sum(1 for c in cleaned if (0 <= ord(c) <= 31) or ord(c) == 127)
                if len(cleaned) > 0 and control_char_count / len(cleaned) > 0.1:
                    logger.warning(f"Detected high percentage of control characters ({control_char_count}/{len(cleaned)}), replacing with empty string")
                    return ''
            # 如果清理后为空，返回空字符串
            if not cleaned:
                return ''
            return cleaned
        
        # 清理列名
        df.columns = [clean_excel_string(str(col)) or f'Column_{i}' for i, col in enumerate(df.columns)]
        
        # 清理所有字符串列的数据
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(lambda x: clean_excel_string(str(x)) if pd.notna(x) and str(x) != 'nan' else '')
        
        # 确保没有空列名
        df.columns = [f'Column_{i}' if not col or col.strip() == '' else col for i, col in enumerate(df.columns)]
        
        # 写入 Excel 文件
        logger.debug(f"Writing Excel file ({output_format.upper()}): {output_path}")
        try:
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 如果输出文件已存在，先删除（避免写入冲突）
            if output_path.exists():
                try:
                    output_path.unlink()
                except Exception as e:
                    logger.warning(f"Could not remove existing file {output_path}: {str(e)}")
            
            # 根据格式选择引擎
            if output_format.lower() == 'xlsx':
                # 使用 openpyxl 引擎写入 XLSX
                try:
                    df.to_excel(output_path, index=False, engine='openpyxl')
                except Exception as e:
                    # 如果写入失败，尝试使用 ExcelWriter 并设置选项
                    from openpyxl import Workbook
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    try:
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Sheet1"
                        
                        # 写入数据
                        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                            for c_idx, value in enumerate(row, 1):
                                # 处理 None 和 NaN 值
                                if pd.isna(value) or value is None:
                                    value = ''
                                # 处理字符串值
                                elif isinstance(value, str):
                                    # 再次清理非法字符（双重保险）
                                    value = clean_excel_string(value)
                                    # 如果清理后为空，使用空字符串
                                    if not value:
                                        value = ''
                                    # 处理过长的字符串（Excel 限制）
                                    elif len(value) > 32767:
                                        value = value[:32767]
                                        logger.warning(f"Truncated long string in row {r_idx}, column {c_idx}")
                                # 处理其他类型（数字等）
                                else:
                                    # 对于非字符串类型，也尝试转换为字符串并清理（以防万一）
                                    if not isinstance(value, (int, float, bool)) and value is not None:
                                        value_str = str(value)
                                        # 检查是否包含非法字符或二进制数据
                                        if re.search(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', value_str):
                                            value = clean_excel_string(value_str)
                                            if not value:
                                                value = ''
                                
                                # 写入单元格
                                try:
                                    ws.cell(row=r_idx, column=c_idx, value=value)
                                except Exception as cell_error:
                                    # 如果单个单元格写入失败，记录并跳过或使用空值
                                    error_msg_short = str(cell_error)[:200] if len(str(cell_error)) > 200 else str(cell_error)
                                    logger.warning(f"Failed to write cell at row {r_idx}, column {c_idx}: {error_msg_short}. Using empty value.")
                                    try:
                                        ws.cell(row=r_idx, column=c_idx, value='')
                                    except:
                                        # 如果连空值都写不进去，跳过这个单元格
                                        pass
                        
                        wb.save(output_path)
                    except Exception as e2:
                        error_msg = f"Failed to write Excel file using openpyxl: {str(e)}. Alternative method also failed: {str(e2)}"
                        logger.error(error_msg, exc_info=True)
                        raise IOError(error_msg) from e2
            else:
                # 使用 xlwt 引擎写入 XLS（需要安装 xlwt）
                try:
                    df.to_excel(output_path, index=False, engine='xlwt')
                except ImportError:
                    # 如果 xlwt 不可用，尝试使用 openpyxl 但会失败，提供清晰的错误信息
                    raise ImportError(
                        "Writing XLS format requires 'xlwt' package. "
                        "Please install it using: pip install xlwt or conda install xlwt. "
                        "Alternatively, use XLSX format which only requires 'openpyxl'."
                    )
            
            # 验证文件是否成功写入
            if not output_path.exists():
                error_msg = f"Excel file was not created: {output_path}"
                logger.error(error_msg)
                raise IOError(error_msg)
            
            if output_path.stat().st_size == 0:
                error_msg = f"Excel file is empty: {output_path}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
        except ImportError as e:
            # 依赖缺失错误，直接抛出
            raise
        except Exception as e:
            error_msg = f"Failed to write Excel file to {output_path}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            # 记录更多调试信息
            logger.debug(f"DataFrame shape: {df.shape}, columns: {list(df.columns)}")
            logger.debug(f"DataFrame dtypes: {df.dtypes.to_dict()}")
            raise IOError(error_msg) from e
        
        conversion_time = time.time() - start_time
        file_size = output_path.stat().st_size / 1024  # KB
        logger.info(f"CSV to Excel ({output_format.upper()}) conversion completed: {len(df)} rows, {len(df.columns)} columns, "
                   f"output size: {file_size:.2f} KB, elapsed: {conversion_time:.2f}s")
        
    except ImportError as e:
        # 依赖缺失错误，直接抛出并提供清晰的错误消息
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'xlwt' in error_str.lower():
            excel_available, excel_error_msg = _check_excel_dependencies()
            if not excel_available:
                error_msg = f"Excel file writing requires dependency. {excel_error_msg}"
                logger.error(error_msg)
                raise ImportError(error_msg) from e
        raise
    except ValueError as e:
        logger.error(f"Value error during CSV to Excel conversion: {str(e)}", exc_info=True)
        raise
    except Exception as e:
        error_str = str(e)
        # 检查是否是依赖缺失的错误
        if 'openpyxl' in error_str.lower() or 'xlwt' in error_str.lower() or 'Missing optional dependency' in error_str:
            excel_available, excel_error_msg = _check_excel_dependencies()
            if not excel_available:
                error_msg = f"Excel file writing requires dependency. {excel_error_msg}"
                logger.error(error_msg)
                raise ImportError(error_msg) from e
        error_msg = f"Unexpected error during CSV to Excel conversion: {str(e)}"
        logger.error(error_msg, exc_info=True)
        raise Exception(error_msg) from e


def _merge_csv_cartesian(antibody_file: Union[str, Path], antigen_file: Union[str, Path], 
                         output_file: Union[str, Path], 
                         key_column1: Optional[str] = None, 
                         key_column2: Optional[str] = None,
                         new_key_name: str = "merged_key") -> None:
    """
    将抗体 CSV 文件和抗原 CSV 文件进行笛卡尔积合并
    
    抗体文件的每一行都与抗原文件的每一行组合，生成所有可能的抗体-抗原对。
    结果行数 = 抗体文件行数 × 抗原文件行数
    两个文件的表头都会保留在结果中，如果列名冲突，会自动添加后缀。
    如果提供了 key_column1 和 key_column2，将创建一个新的 key 列，由这两列的值拼接而成。
    
    Args:
        antibody_file: 抗体 CSV 文件路径（必须包含 Heavy 和 Light 列）
        antigen_file: 抗原 CSV 文件路径（必须包含 variant_seq、antigen_seq 或 variant 列）
        output_file: 输出的合并 CSV 文件路径
        key_column1: 来自抗体文件的列名，用于生成新 key（可选）
        key_column2: 来自抗原文件的列名，用于生成新 key（可选）
        new_key_name: 新生成的 key 列的名称（默认: "merged_key"）
        
    Raises:
        ValueError: 如果文件验证失败、缺少必需列或数据格式错误
        Exception: 如果文件读取或写入失败
    """
    antibody_path = Path(antibody_file)
    antigen_path = Path(antigen_file)
    output_path = Path(output_file)
    
    logger.info(f"Starting antibody-antigen cartesian merge: {antibody_path} × {antigen_path} -> {output_path}")
    start_time = time.time()
    
    # 验证两个 CSV 文件
    is_valid1, error_msg1 = _validate_csv_file(antibody_path)
    if not is_valid1:
        logger.error(f"Antibody CSV file validation failed: {error_msg1}")
        raise ValueError(f"Antibody CSV file validation failed: {error_msg1}")
    
    is_valid2, error_msg2 = _validate_csv_file(antigen_path)
    if not is_valid2:
        logger.error(f"Antigen CSV file validation failed: {error_msg2}")
        raise ValueError(f"Antigen CSV file validation failed: {error_msg2}")
    
    logger.info("Both CSV files validation passed")
    
    try:
        # 读取两个 CSV 文件
        antibody_df = _read_csv_robust(antibody_path)
        antigen_df = _read_csv_robust(antigen_path)
        
        logger.info(f"Loaded antibody file: {len(antibody_df)} rows, {len(antibody_df.columns)} columns")
        logger.info(f"Loaded antigen file: {len(antigen_df)} rows, {len(antigen_df.columns)} columns")
        
        # 检查数据框是否为空
        if antibody_df.empty:
            raise ValueError("Antibody CSV file is empty")
        if antigen_df.empty:
            raise ValueError("Antigen CSV file is empty")
        
        # 处理列名冲突：为列名添加后缀
        antibody_columns = antibody_df.columns.tolist()
        antigen_columns = antigen_df.columns.tolist()
        
        # 找出冲突的列名
        common_columns = set(antibody_columns) & set(antigen_columns)
        
        # 重命名冲突的列
        if common_columns:
            logger.info(f"Found {len(common_columns)} common column names, adding suffixes")
            rename_dict_ab = {col: f"{col}_antibody" for col in common_columns}
            rename_dict_ag = {col: f"{col}_antigen" for col in common_columns}
            antibody_df = antibody_df.rename(columns=rename_dict_ab)
            antigen_df = antigen_df.rename(columns=rename_dict_ag)
        
        # 执行笛卡尔积合并
        # 为每个 DataFrame 添加一个临时键列用于合并
        antibody_df['_temp_key'] = 1
        antigen_df['_temp_key'] = 1
        
        # 执行笛卡尔积（使用 inner join 在临时键上）
        merged_df = antibody_df.merge(antigen_df, on='_temp_key', how='inner')
        
        # 删除临时键列
        merged_df = merged_df.drop(columns=['_temp_key'])
        
        # 如果提供了 key_column1 和 key_column2，生成新的 key 列
        if key_column1 and key_column2:
            # 检查列是否存在（考虑重命名后的列名）
            actual_key_col1 = key_column1
            actual_key_col2 = key_column2
            
            # 如果列被重命名了，需要找到实际列名
            if key_column1 in common_columns:
                actual_key_col1 = f"{key_column1}_antibody"
            if key_column2 in common_columns:
                actual_key_col2 = f"{key_column2}_antigen"
            
            # 验证列是否存在（在合并后的数据框中检查）
            if actual_key_col1 not in merged_df.columns:
                # 如果原始列名也不存在，说明列根本不存在
                if key_column1 not in merged_df.columns:
                    raise ValueError(f"Key column '{key_column1}' not found in antibody file. Available columns: {list(antibody_df.columns)}")
                else:
                    actual_key_col1 = key_column1
            
            if actual_key_col2 not in merged_df.columns:
                # 如果原始列名也不存在，说明列根本不存在
                if key_column2 not in merged_df.columns:
                    raise ValueError(f"Key column '{key_column2}' not found in antigen file. Available columns: {list(antigen_df.columns)}")
                else:
                    actual_key_col2 = key_column2
            
            # 生成新的 key 列（拼接两列的值）
            merged_df[new_key_name] = merged_df[actual_key_col1].astype(str) + "_" + merged_df[actual_key_col2].astype(str)
            
            # 将新 key 列移到第一位
            cols = [new_key_name] + [col for col in merged_df.columns if col != new_key_name]
            merged_df = merged_df[cols]
            
            logger.info(f"Created new key column '{new_key_name}' by concatenating '{key_column1}' (as '{actual_key_col1}') and '{key_column2}' (as '{actual_key_col2}')")
        
        logger.info(f"Antibody-antigen cartesian merge completed: {len(merged_df)} rows ({len(antibody_df)} antibodies × {len(antigen_df)} antigens)")
        
        # 写入合并后的 CSV 文件
        logger.debug(f"Writing merged CSV file: {output_path}")
        try:
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 保存为 CSV
            merged_df.to_csv(output_path, index=False)
            
            # 验证文件是否成功写入
            if not output_path.exists():
                error_msg = f"Merged CSV file was not created: {output_path}"
                logger.error(error_msg)
                raise IOError(error_msg)
            
            if output_path.stat().st_size == 0:
                error_msg = f"Merged CSV file is empty: {output_path}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
        except Exception as e:
            error_msg = f"Failed to write merged CSV file to {output_path}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise IOError(error_msg) from e
        
        merge_time = time.time() - start_time
        file_size = output_path.stat().st_size / 1024  # KB
        logger.info(f"CSV cartesian merge completed: {len(merged_df)} rows, "
                   f"output size: {file_size:.2f} KB, elapsed: {merge_time:.2f}s")
        
    except ValueError as e:
        logger.error(f"Value error during CSV cartesian merge: {str(e)}", exc_info=True)
        raise
    except Exception as e:
        error_msg = f"Unexpected error during CSV cartesian merge: {str(e)}"
        logger.error(error_msg, exc_info=True)
        raise Exception(error_msg) from e


def _merge_csv_by_key(input_file1: Union[str, Path], input_file2: Union[str, Path], 
                      output_file: Union[str, Path], key_column: Optional[str] = None) -> None:
    """
    合并两个 CSV 文件（加法运算）
    
    如果提供了 key_column，则基于指定 key 列合并，使用并集（outer join）。
    如果没有提供 key_column，则按照行索引一一对应进行拼接。
    表头取并集，如果列名冲突，直接使用第一个文件的值（不添加后缀）。
    如果其中一个文件为空，则直接返回另一个非空文件。
    
    Args:
        input_file1: 第一个 CSV 文件路径
        input_file2: 第二个 CSV 文件路径
        output_file: 输出的合并 CSV 文件路径
        key_column: 可选的用于合并的 key 列名（如果提供，必须在两个文件中都存在，除非其中一个文件为空）
        
    Raises:
        ValueError: 如果文件验证失败、行数不一致、key 列不存在或数据格式错误
        Exception: 如果文件读取或写入失败
    """
    input_path1 = Path(input_file1)
    input_path2 = Path(input_file2)
    output_path = Path(output_file)
    
    if key_column:
        logger.info(f"Starting CSV merge by key '{key_column}': {input_path1} + {input_path2} -> {output_path}")
    else:
        logger.info(f"Starting CSV merge by row index: {input_path1} + {input_path2} -> {output_path}")
    start_time = time.time()
    
    # 验证两个 CSV 文件（允许空文件）
    is_valid1, error_msg1 = _validate_csv_file(input_path1)
    if not is_valid1:
        # 检查是否是因为文件为空
        if input_path1.exists() and input_path1.stat().st_size == 0:
            logger.warning(f"CSV file 1 is empty, will return file 2 if it's not empty")
        else:
            logger.error(f"CSV file 1 validation failed: {error_msg1}")
            raise ValueError(f"CSV file 1 validation failed: {error_msg1}")
    
    is_valid2, error_msg2 = _validate_csv_file(input_path2)
    if not is_valid2:
        # 检查是否是因为文件为空
        if input_path2.exists() and input_path2.stat().st_size == 0:
            logger.warning(f"CSV file 2 is empty, will return file 1 if it's not empty")
        else:
            logger.error(f"CSV file 2 validation failed: {error_msg2}")
            raise ValueError(f"CSV file 2 validation failed: {error_msg2}")
    
    logger.info("CSV files validation passed")
    
    try:
        # 读取两个 CSV 文件（允许空文件）
        try:
            df1 = _read_csv_robust(input_path1)
        except Exception as e:
            # 如果读取失败且文件为空，创建空 DataFrame
            if input_path1.exists() and input_path1.stat().st_size == 0:
                logger.info("File 1 is empty, creating empty DataFrame")
                df1 = pd.DataFrame()
            else:
                raise
        
        try:
            df2 = _read_csv_robust(input_path2)
        except Exception as e:
            # 如果读取失败且文件为空，创建空 DataFrame
            if input_path2.exists() and input_path2.stat().st_size == 0:
                logger.info("File 2 is empty, creating empty DataFrame")
                df2 = pd.DataFrame()
            else:
                raise
        
        logger.info(f"Loaded file 1: {len(df1)} rows, {len(df1.columns)} columns")
        logger.info(f"Loaded file 2: {len(df2)} rows, {len(df2.columns)} columns")
        
        # 处理空文件的情况：如果其中一个为空，直接返回另一个
        if df1.empty and not df2.empty:
            logger.info("File 1 is empty, returning file 2 as output")
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            # 复制文件2到输出路径
            import shutil
            shutil.copy2(input_path2, output_path)
            logger.info(f"Copied file 2 to output: {output_path}")
            return
        
        if df2.empty and not df1.empty:
            logger.info("File 2 is empty, returning file 1 as output")
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            # 复制文件1到输出路径
            import shutil
            shutil.copy2(input_path1, output_path)
            logger.info(f"Copied file 1 to output: {output_path}")
            return
        
        if df1.empty and df2.empty:
            logger.warning("Both files are empty, creating empty output file")
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            # 创建空文件
            with open(output_path, 'w', encoding='utf-8') as f:
                pass
            logger.info(f"Created empty output file: {output_path}")
            return
        
        # 根据是否提供 key_column 选择合并方式
        # 注意：使用并集合并时，行数可以不同
        if key_column:
            # 基于 key 列合并
            # 验证 key 列是否存在
            if key_column not in df1.columns:
                error_msg = f"Key column '{key_column}' not found in file 1. Available columns: {list(df1.columns)}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            if key_column not in df2.columns:
                error_msg = f"Key column '{key_column}' not found in file 2. Available columns: {list(df2.columns)}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            # 验证 key 值是否一一对应
            key1 = df1[key_column].astype(str).values
            key2 = df2[key_column].astype(str).values
            
            if not all(k1 == k2 for k1, k2 in zip(key1, key2)):
                error_msg = f"Key values do not match between the two files. " \
                           f"Key values must be in the same order and match exactly."
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            logger.info(f"Key column '{key_column}' validation passed: {len(key1)} matching keys")
            
            # 处理列名冲突：除了 key 列，其他冲突的列名直接使用第一个文件的值
            df1_columns = [col for col in df1.columns if col != key_column]
            df2_columns = [col for col in df2.columns if col != key_column]
            
            # 找出冲突的列名（不包括 key 列）
            common_columns = set(df1_columns) & set(df2_columns)
            
            # 对于冲突的列，从第二个文件中删除，保留第一个文件的值
            if common_columns:
                logger.info(f"Found {len(common_columns)} common column names (excluding key), using values from file 1")
                # 从df2中删除冲突的列（除了key列）
                df2 = df2.drop(columns=[col for col in common_columns if col in df2.columns])
            
            # 基于 key 列合并，使用并集（outer join）
            merged_df = df1.merge(df2, on=key_column, how='outer')
            
            # 确保 key 列在第一位
            cols = [key_column] + [col for col in merged_df.columns if col != key_column]
            merged_df = merged_df[cols]
            
            logger.info(f"Key-based merge completed: {len(merged_df)} rows, {len(merged_df.columns)} columns")
        else:
            # 按行索引一一对应拼接
            logger.info("Merging files by row index (row-by-row concatenation)")
            
            # 处理列名冲突：所有冲突的列名直接使用第一个文件的值
            df1_columns = list(df1.columns)
            df2_columns = list(df2.columns)
            
            # 找出冲突的列名
            common_columns = set(df1_columns) & set(df2_columns)
            
            # 对于冲突的列，从第二个文件中删除，保留第一个文件的值
            if common_columns:
                logger.info(f"Found {len(common_columns)} common column names, using values from file 1")
                # 从df2中删除冲突的列
                df2 = df2.drop(columns=[col for col in common_columns if col in df2.columns])
            
            # 重置索引以确保按行一一对应
            df1 = df1.reset_index(drop=True)
            df2 = df2.reset_index(drop=True)
            
            # 按列拼接（axis=1 表示按列拼接，即行对行）
            # 取并集：如果行数不同，用NaN填充
            max_rows = max(len(df1), len(df2))
            if len(df1) < max_rows:
                # 扩展df1到最大行数，用NaN填充
                df1 = df1.reindex(range(max_rows))
            if len(df2) < max_rows:
                # 扩展df2到最大行数，用NaN填充
                df2 = df2.reindex(range(max_rows))
            
            merged_df = pd.concat([df1, df2], axis=1)
            
            logger.info(f"Row-index-based merge completed: {len(merged_df)} rows, {len(merged_df.columns)} columns")
        
        # 写入合并后的 CSV 文件
        logger.debug(f"Writing merged CSV file: {output_path}")
        try:
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # 保存为 CSV
            merged_df.to_csv(output_path, index=False)
            
            # 验证文件是否成功写入
            if not output_path.exists():
                error_msg = f"Merged CSV file was not created: {output_path}"
                logger.error(error_msg)
                raise IOError(error_msg)
            
            if output_path.stat().st_size == 0:
                error_msg = f"Merged CSV file is empty: {output_path}"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
        except Exception as e:
            error_msg = f"Failed to write merged CSV file to {output_path}: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise IOError(error_msg) from e
        
        merge_time = time.time() - start_time
        file_size = output_path.stat().st_size / 1024  # KB
        logger.info(f"CSV key-based merge completed: {len(merged_df)} rows, "
                   f"output size: {file_size:.2f} KB, elapsed: {merge_time:.2f}s")
        
    except ValueError as e:
        logger.error(f"Value error during CSV key-based merge: {str(e)}", exc_info=True)
        raise
    except Exception as e:
        error_msg = f"Unexpected error during CSV key-based merge: {str(e)}"
        logger.error(error_msg, exc_info=True)
        raise Exception(error_msg) from e


# ==================== MCP 工具定义 ====================

class DownloadUrlArgs(BaseModel):
    """Parameters for downloading file from URL"""
    url: str = Field(
        ...,
        description="HTTP/HTTPS URL to download file from",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "https://example.com/file.fasta",
            "help_text": "Enter HTTP or HTTPS URL to download file. The file will be saved to a temporary location and the path will be returned."
        }
    )
    output_path: Optional[str] = Field(
        default=None,
        description="Optional output file path. If not provided, a temporary file will be created automatically.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/output/file.fasta (optional)",
            "help_text": "Optional: Specify output file path. If not provided, a temporary file will be created."
        }
    )


class ConvertCsvToFastaArgs(BaseModel):
    """Parameters for converting CSV to FASTA format"""
    input_file: str = Field(
        ...,
        description="Input CSV file path or HTTP/HTTPS URL to CSV file",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["csv"],
            "placeholder": "Enter CSV file path or URL (e.g., /path/to/file.csv or https://example.com/file.csv)",
            "help_text": "Input file in CSV format. Can be local file path or HTTP/HTTPS URL. The file should contain ID and sequence columns (auto-detected)."
        }
    )
    output_file: Optional[str] = Field(
        default=None,
        description="Output FASTA file path. If not provided, will be auto-generated based on input file name.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/output.fasta (optional)",
            "help_text": "Optional: Specify output FASTA file path. If not provided, will be auto-generated in the same directory as input file."
        }
    )


class ConvertExcelToFastaArgs(BaseModel):
    """Parameters for converting Excel to FASTA format"""
    input_file: str = Field(
        ...,
        description="Input Excel file path (XLSX or XLS) or HTTP/HTTPS URL to Excel file",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["xlsx", "xls"],
            "placeholder": "Enter Excel file path or URL (e.g., /path/to/file.xlsx or https://example.com/file.xlsx)",
            "help_text": "Input file in Excel format (XLSX or XLS). Can be local file path or HTTP/HTTPS URL. The file should contain ID and sequence columns (auto-detected)."
        }
    )
    output_file: Optional[str] = Field(
        default=None,
        description="Output FASTA file path. If not provided, will be auto-generated based on input file name.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/output.fasta (optional)",
            "help_text": "Optional: Specify output FASTA file path. If not provided, will be auto-generated in the same directory as input file."
        }
    )


class ConvertXlsxToFastaArgs(BaseModel):
    """Parameters for converting XLSX to FASTA format"""
    input_file: str = Field(
        ...,
        description="Input XLSX file path or HTTP/HTTPS URL to XLSX file",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["xlsx"],
            "placeholder": "Enter XLSX file path or URL (e.g., /path/to/file.xlsx or https://example.com/file.xlsx)",
            "help_text": "Input file in XLSX format. Can be local file path or HTTP/HTTPS URL. The file should contain ID and sequence columns (auto-detected)."
        }
    )
    output_file: Optional[str] = Field(
        default=None,
        description="Output FASTA file path. If not provided, will be auto-generated based on input file name.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/output.fasta (optional)",
            "help_text": "Optional: Specify output FASTA file path. If not provided, will be auto-generated in the same directory as input file."
        }
    )


class ConvertXlsToFastaArgs(BaseModel):
    """Parameters for converting XLS to FASTA format"""
    input_file: str = Field(
        ...,
        description="Input XLS file path or HTTP/HTTPS URL to XLS file",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["xls"],
            "placeholder": "Enter XLS file path or URL (e.g., /path/to/file.xls or https://example.com/file.xls)",
            "help_text": "Input file in XLS format (Excel 97-2003). Can be local file path or HTTP/HTTPS URL. The file should contain ID and sequence columns (auto-detected)."
        }
    )
    output_file: Optional[str] = Field(
        default=None,
        description="Output FASTA file path. If not provided, will be auto-generated based on input file name.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/output.fasta (optional)",
            "help_text": "Optional: Specify output FASTA file path. If not provided, will be auto-generated in the same directory as input file."
        }
    )


class ConvertXlsxToCsvArgs(BaseModel):
    """Parameters for converting XLSX to CSV format"""
    input_file: str = Field(
        ...,
        description="Input XLSX file path or HTTP/HTTPS URL to XLSX file",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["xlsx"],
            "placeholder": "Enter XLSX file path or URL (e.g., /path/to/file.xlsx or https://example.com/file.xlsx)",
            "help_text": "Input file in XLSX format. Can be local file path or HTTP/HTTPS URL. All columns and data will be preserved in the output CSV."
        }
    )
    output_file: Optional[str] = Field(
        default=None,
        description="Output CSV file path. If not provided, will be auto-generated based on input file name.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/output.csv (optional)",
            "help_text": "Optional: Specify output CSV file path. If not provided, will be auto-generated in the same directory as input file."
        }
    )


class ConvertXlsToCsvArgs(BaseModel):
    """Parameters for converting XLS to CSV format"""
    input_file: str = Field(
        ...,
        description="Input XLS file path or HTTP/HTTPS URL to XLS file",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["xls"],
            "placeholder": "Enter XLS file path or URL (e.g., /path/to/file.xls or https://example.com/file.xls)",
            "help_text": "Input file in XLS format (Excel 97-2003). Can be local file path or HTTP/HTTPS URL. All columns and data will be preserved in the output CSV."
        }
    )
    output_file: Optional[str] = Field(
        default=None,
        description="Output CSV file path. If not provided, will be auto-generated based on input file name.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/output.csv (optional)",
            "help_text": "Optional: Specify output CSV file path. If not provided, will be auto-generated in the same directory as input file."
        }
    )


class ConvertCsvToXlsxArgs(BaseModel):
    """Parameters for converting CSV to XLSX format"""
    input_file: str = Field(
        ...,
        description="Input CSV file path or HTTP/HTTPS URL to CSV file",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["csv"],
            "placeholder": "Enter CSV file path or URL (e.g., /path/to/file.csv or https://example.com/file.csv)",
            "help_text": "Input file in CSV format. Can be local file path or HTTP/HTTPS URL. All columns and data will be preserved in the output Excel file."
        }
    )
    output_file: Optional[str] = Field(
        default=None,
        description="Output XLSX file path. If not provided, will be auto-generated based on input file name.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/output.xlsx (optional)",
            "help_text": "Optional: Specify output XLSX file path. If not provided, will be auto-generated in the same directory as input file."
        }
    )


class CreateCsvArgs(BaseModel):
    """Parameters for creating a blank CSV file"""
    output_file: Optional[str] = Field(
        default=None,
        description="Output CSV file path. If not provided, a temporary file will be created automatically.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/output.csv (optional)",
            "help_text": "Optional: Specify output CSV file path. If not provided, a temporary file will be created and the path will be returned."
        }
    )
    columns: Optional[List[str]] = Field(
        default=None,
        description="Optional list of column names to include as header in the CSV file. If not provided, an empty CSV file will be created.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "Enter comma-separated column names (e.g., 'id,name,sequence')",
            "help_text": "Optional: Comma-separated list of column names. If provided, these will be written as the header row in the CSV file."
        }
    )


class MergeCsvCartesianArgs(BaseModel):
    """Parameters for cartesian merge of antibody and antigen CSV files"""
    antibody_file: str = Field(
        ...,
        description="Antibody CSV file path or HTTP/HTTPS URL",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["csv"],
            "placeholder": "Enter antibody CSV file path or URL",
            "help_text": "Antibody CSV file for cartesian merge. Must contain 'Heavy' and 'Light' columns. Each antibody will be combined with each antigen."
        }
    )
    antigen_file: str = Field(
        ...,
        description="Antigen CSV file path or HTTP/HTTPS URL",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["csv"],
            "placeholder": "Enter antigen CSV file path or URL",
            "help_text": "Antigen CSV file for cartesian merge. Must contain 'variant_seq' (or 'antigen_seq' or 'variant') column. Each antigen will be combined with each antibody."
        }
    )
    antibody_key: Optional[str] = Field(
        default=None,
        description="Column name from antibody file to use for generating new key. If provided together with key_column2, a new key column will be created by concatenating values from these two columns.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "Enter column name from antibody file (e.g., 'barcode', 'id')",
            "help_text": "Optional: Column name from the antibody file. If both key_column1 and key_column2 are provided, a new key column will be created by concatenating their values."
        }
    )
    antigen_key: Optional[str] = Field(
        default=None,
        description="Column name from antigen file to use for generating new key. If provided together with key_column1, a new key column will be created by concatenating values from these two columns.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "Enter column name from antigen file (e.g., 'variant_id', 'antigen_id')",
            "help_text": "Optional: Column name from the antigen file. If both key_column1 and key_column2 are provided, a new key column will be created by concatenating their values."
        }
    )
    new_key_name: Optional[str] = Field(
        default="merged_key",
        description="Name for the new key column created by concatenating key_column1 and key_column2.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "Enter new key column name (default: 'merged_key')",
            "help_text": "Optional: Name for the new key column. Default is 'merged_key'. Only used if both key_column1 and key_column2 are provided."
        }
    )
    output_file: Optional[str] = Field(
        default=None,
        description="Output merged CSV file path. If not provided, will be auto-generated.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/merged_output.csv (optional)",
            "help_text": "Optional: Specify output CSV file path. If not provided, will be auto-generated."
        }
    )


class MergeCsvByKeyArgs(BaseModel):
    """Parameters for key-based merge of two CSV files"""
    input_file1: str = Field(
        ...,
        description="First CSV file path or HTTP/HTTPS URL",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["csv"],
            "placeholder": "Enter first CSV file path or URL",
            "help_text": "First CSV file for key-based merge. Must have the same number of rows as file 2."
        }
    )
    input_file2: str = Field(
        ...,
        description="Second CSV file path or HTTP/HTTPS URL",
        json_schema_extra={
            "ui_type": "file_input",
            "support_upload": True,
            "support_file_types": ["csv"],
            "placeholder": "Enter second CSV file path or URL",
            "help_text": "Second CSV file for key-based merge. Must have the same number of rows as file 1."
        }
    )
    key_column: Optional[str] = Field(
        default=None,
        description="Optional column name to use as merge key (must exist in both files). If not provided, files will be merged by row index (row-by-row concatenation).",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "Enter key column name (e.g., 'id', 'sample_id') or leave empty for row-by-row merge",
            "help_text": "Optional: Column name that exists in both CSV files. If not provided, files will be merged row-by-row based on their position. Key values must match in the same order if provided."
        }
    )
    output_file: Optional[str] = Field(
        default=None,
        description="Output merged CSV file path. If not provided, will be auto-generated.",
        json_schema_extra={
            "ui_type": "text_input",
            "placeholder": "/path/to/merged_output.csv (optional)",
            "help_text": "Optional: Specify output CSV file path. If not provided, will be auto-generated."
        }
    )


@mcp.tool()
async def download_url(args: DownloadUrlArgs):
    """
    从 HTTP/HTTPS URL 下载文件到本地
    
    支持下载任意类型的文件，自动根据 URL 中的扩展名确定文件类型。
    如果未指定输出路径，将创建临时文件。
    
    通过 SSE 流式推送下载进度，支持大文件下载。
    
    Args:
        args: DownloadUrlArgs - 下载参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    start_time = time.time()
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting file download from URL: {args.url}")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": f"Starting download from {args.url}",
            "timestamp": time.time()
        }
    }
    
    temp_file_path = None
    
    try:
        # 如果提供了输出路径，使用它；否则创建临时文件
        if args.output_path:
            output_path = Path(args.output_path)
            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)
            temp_file_path = str(output_path)
            logger.info(f"[{session_id}] Using specified output path: {temp_file_path}")
        else:
            # 生成临时文件路径
            url_path = urlparse(args.url).path
            file_ext = os.path.splitext(url_path)[1] or '.tmp'
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=file_ext)
            temp_file_path = temp_file.name
            temp_file.close()
            logger.info(f"[{session_id}] Created temporary file: {temp_file_path}")
        
        # 发送开始下载进度
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "processing",
                "message": "Connecting to server...",
                "timestamp": time.time()
            }
        }
        
        # 使用改进的下载方式（以二进制模式下载，确保编码正确）
        if HAS_REQUESTS:
            response = requests.get(args.url, stream=True, timeout=30)
            response.raise_for_status()
            
            # 获取文件大小（如果可用）
            total_size = int(response.headers.get('content-length', 0))
            
            # 记录响应头信息（用于调试）
            content_type = response.headers.get('Content-Type', '')
            content_encoding = response.headers.get('Content-Encoding', '')
            logger.debug(f"[{session_id}] Response headers - Content-Type: {content_type}, Content-Encoding: {content_encoding}")
            logger.info(f"[{session_id}] File size: {total_size / (1024 * 1024):.2f} MB" if total_size > 0 else f"[{session_id}] File size: unknown")
            
            # 以二进制模式写入文件（确保不进行任何编码转换）
            downloaded = 0
            last_progress_time = time.time()
            last_progress_bytes = 0
            progress_interval = 30.0  # 每30秒推送一次进度
            progress_bytes_interval = 5 * 1024 * 1024  # 每5MB推送一次进度
            
            with open(temp_file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        f.write(chunk)
                        downloaded += len(chunk)
                        
                        # 推送进度（每5MB或每30秒推送一次，避免过于频繁）
                        current_time = time.time()
                        bytes_since_last = downloaded - last_progress_bytes
                        time_since_last = current_time - last_progress_time
                        should_push = (
                            bytes_since_last >= progress_bytes_interval or  # 每5MB
                            time_since_last >= progress_interval or  # 每30秒
                            (total_size > 0 and downloaded >= total_size)  # 下载完成
                        )
                        
                        if should_push:
                            progress_percent = (downloaded / total_size * 100) if total_size > 0 else 0
                            elapsed_time = current_time - start_time
                            download_speed = downloaded / elapsed_time if elapsed_time > 0 else 0
                            
                            # 计算ETA
                            if total_size > 0 and download_speed > 0 and downloaded < total_size:
                                remaining = (total_size - downloaded) / download_speed
                            else:
                                remaining = None
                            
                            yield {
                                "type": "progress",
                                "data": {
                                    "session_id": session_id,
                                    "status": "processing",
                                    "progress_percent": round(progress_percent, 1) if total_size > 0 else None,
                                    "downloaded_bytes": downloaded,
                                    "downloaded_mb": round(downloaded / (1024 * 1024), 2),
                                    "total_size_bytes": total_size if total_size > 0 else None,
                                    "total_size_mb": round(total_size / (1024 * 1024), 2) if total_size > 0 else None,
                                    "download_speed_mbps": round(download_speed / (1024 * 1024), 2) if download_speed > 0 else None,
                                    "elapsed_seconds": round(elapsed_time, 1),
                                    "eta_seconds": round(remaining, 1) if remaining else None,
                                    "eta_minutes": round(remaining / 60, 1) if remaining else None,
                                    "message": f"Downloading: {downloaded / (1024 * 1024):.2f} MB" + (f" / {total_size / (1024 * 1024):.2f} MB ({progress_percent:.1f}%)" if total_size > 0 else ""),
                                    "timestamp": current_time
                                }
                            }
                            last_progress_time = current_time
                            last_progress_bytes = downloaded
        else:
            # 回退到 urllib（不支持进度推送，但可以发送开始和完成消息）
            yield {
                "type": "progress",
                "data": {
                    "session_id": session_id,
                    "status": "processing",
                    "message": "Downloading file (progress not available with urllib)...",
                    "timestamp": time.time()
                }
            }
            
            urllib.request.urlretrieve(args.url, temp_file_path)
            
            # urllib 下载完成后发送一次进度
            file_size = os.path.getsize(temp_file_path) / (1024 * 1024)
            yield {
                "type": "progress",
                "data": {
                    "session_id": session_id,
                    "status": "processing",
                    "progress_percent": 100.0,
                    "downloaded_mb": round(file_size, 2),
                    "message": f"Downloaded: {file_size:.2f} MB",
                    "timestamp": time.time()
                }
            }
        
        logger.info(f"[{session_id}] Downloaded file to path: {temp_file_path}")
        
        # 验证文件是否成功下载
        if not os.path.exists(temp_file_path):
            error_msg = f"[{session_id}] Downloaded file does not exist: {temp_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "status": "error",
                "error_type": "file_not_found",
                "message": error_msg,
                "session_id": session_id
            }
            return
        
        file_size = os.path.getsize(temp_file_path) / (1024 * 1024)  # MB
        
        logger.info(f"[{session_id}] File download completed successfully: {temp_file_path} ({file_size:.2f} MB)")
        
        # 发送完成进度
        total_time = time.time() - start_time
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "completed",
                "progress_percent": 100.0,
                "elapsed_seconds": round(total_time, 1),
                "message": "Download completed",
                "timestamp": time.time()
            }
        }
        
        # 发送最终结果
        yield {
            "type": "result",
            "status": "success",
            "session_id": session_id,
            "file_path": temp_file_path,
            "file_size_mb": round(file_size, 2),
            "is_temporary": args.output_path is None,
            "url": args.url,
            "processing_time_ms": total_time * 1000
        }
        
    except requests.exceptions.RequestException as e:
        error_msg = f"[{session_id}] Failed to download file from URL {args.url}: {str(e)}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "status": "error",
            "error_type": "download_failed",
            "message": error_msg,
            "url": args.url,
            "session_id": session_id
        }
        return
    except Exception as e:
        error_msg = f"[{session_id}] Failed to download file from URL {args.url}: {str(e)}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "status": "error",
            "error_type": "download_failed",
            "message": error_msg,
            "url": args.url,
            "session_id": session_id
        }
        return


@mcp.tool()
async def convert_csv_to_fasta(args: ConvertCsvToFastaArgs):
    """
    将 CSV 文件转换为 FASTA 格式
    
    自动检测列名，支持多种常见的列名变体。
    对格式不规范的 CSV 文件具有强大的容错能力。
    支持本地文件路径和 HTTP/HTTPS URL。
    
    通过 SSE 流式推送转换进度，支持大文件转换。
    
    Args:
        args: ConvertCsvToFastaArgs - 转换参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting CSV to FASTA conversion")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting CSV to FASTA conversion",
            "timestamp": time.time()
        }
    }
    
    temp_file_path = None
    input_file_path = None
    
    try:
        # 处理输入文件：可能是 URL 或本地路径
        if args.input_file.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Input is URL, downloading first: {args.input_file}")
            yield {
                "type": "progress",
                "data": {
                    "session_id": session_id,
                    "status": "downloading",
                    "message": f"Downloading input file from {args.input_file}",
                    "timestamp": time.time()
                }
            }
            try:
                temp_file_path = download_to_temp_file(args.input_file)
                input_file_path = Path(temp_file_path)
                logger.info(f"[{session_id}] Downloaded input file to: {input_file_path}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download input file from URL {args.input_file}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            input_file_path = Path(args.input_file)
            if not input_file_path.exists():
                error_msg = f"[{session_id}] Input file not found: {args.input_file}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
        else:
            # 自动生成输出文件名：基于输入文件名，替换扩展名为 .fasta
            output_file_path = input_file_path.with_suffix('.fasta')
            logger.info(f"[{session_id}] Auto-generated output path: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 检测文件实际类型，如果是 Excel 文件则自动使用 Excel 转换
        actual_type = _detect_file_type(input_file_path)
        if actual_type in ['.xlsx', '.xls']:
            logger.info(f"[{session_id}] Detected Excel format ({actual_type}), automatically using Excel converter")
            yield {
                "type": "progress",
                "data": {
                    "session_id": session_id,
                    "status": "converting",
                    "message": f"Converting Excel to FASTA: {input_file_path.name}",
                    "timestamp": time.time()
                }
            }
            # 使用 Excel 转换函数
            _convert_excel_to_fasta(input_file_path, output_file_path)
            conversion_type = "excel_to_fasta"
        else:
            # 执行 CSV 转换
            logger.info(f"[{session_id}] Converting CSV to FASTA: {input_file_path} -> {output_file_path}")
            yield {
                "type": "progress",
                "data": {
                    "session_id": session_id,
                    "status": "converting",
                    "message": f"Converting CSV to FASTA: {input_file_path.name}",
                    "timestamp": time.time()
                }
            }
            _convert_csv_to_fasta(input_file_path, output_file_path)
            conversion_type = "csv_to_fasta"
        
        # 验证输出文件
        if not output_file_path.exists():
            error_msg = f"[{session_id}] Output FASTA file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "conversion_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size / 1024  # KB
        
        # 统计序列数量
        num_sequences = 0
        with open(output_file_path, 'r') as f:
            num_sequences = sum(1 for line in f if line.startswith('>'))
        
        conversion_msg = "Excel to FASTA" if conversion_type == "excel_to_fasta" else "CSV to FASTA"
        logger.info(f"[{session_id}] {conversion_msg} conversion completed: {output_file_path} ({num_sequences} sequences, {file_size:.2f} KB)")
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": {
                "status": "success",
                "output_file": str(output_file_path.absolute()),
                "input_file": str(input_file_path.absolute()),
                "num_sequences": num_sequences,
                "file_size_kb": round(file_size, 2),
                "format": "FASTA",
                "conversion_type": conversion_type
            }
        }
        
    except ValueError as e:
        error_msg = f"[{session_id}] CSV to FASTA conversion failed (ValueError): {str(e)}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "conversion_failed",
                "message": error_msg
            }
        }
    except Exception as e:
        error_msg = f"[{session_id}] Unexpected error during CSV to FASTA conversion: {str(e)}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "unknown",
                "message": error_msg
            }
        }
    finally:
        # 清理从 URL 下载的临时输入文件
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.debug(f"[{session_id}] Cleaned up temporary input file: {temp_file_path}")
            except:
                pass


@mcp.tool()
async def convert_excel_to_fasta(args: ConvertExcelToFastaArgs):
    """
    将 Excel 文件转换为 FASTA 格式
    
    自动检测列名，支持多种常见的列名变体。
    支持 XLSX 和 XLS 格式。
    支持本地文件路径和 HTTP/HTTPS URL。
    
    通过 SSE 流式推送转换进度，支持大文件转换。
    
    Args:
        args: ConvertExcelToFastaArgs - 转换参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting Excel to FASTA conversion")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting Excel to FASTA conversion",
            "timestamp": time.time()
        }
    }
    
    temp_file_path = None
    input_file_path = None
    
    try:
        # 处理输入文件：可能是 URL 或本地路径
        if args.input_file.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Input is URL, downloading first: {args.input_file}")
            yield {
                "type": "progress",
                "data": {
                    "session_id": session_id,
                    "status": "downloading",
                    "message": f"Downloading input file from {args.input_file}",
                    "timestamp": time.time()
                }
            }
            try:
                temp_file_path = download_to_temp_file(args.input_file)
                input_file_path = Path(temp_file_path)
                logger.info(f"[{session_id}] Downloaded input file to: {input_file_path}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download input file from URL {args.input_file}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            input_file_path = Path(args.input_file)
            if not input_file_path.exists():
                error_msg = f"[{session_id}] Input file not found: {args.input_file}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
        else:
            # 自动生成输出文件名：基于输入文件名，替换扩展名为 .fasta
            output_file_path = input_file_path.with_suffix('.fasta')
            logger.info(f"[{session_id}] Auto-generated output path: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 执行转换
        logger.info(f"[{session_id}] Converting Excel to FASTA: {input_file_path} -> {output_file_path}")
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "converting",
                "message": f"Converting Excel to FASTA: {input_file_path.name}",
                "timestamp": time.time()
            }
        }
        _convert_excel_to_fasta(input_file_path, output_file_path)
        
        # 验证输出文件
        if not output_file_path.exists():
            error_msg = f"[{session_id}] Output FASTA file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "conversion_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size / 1024  # KB
        
        # 统计序列数量
        num_sequences = 0
        with open(output_file_path, 'r') as f:
            num_sequences = sum(1 for line in f if line.startswith('>'))
        
        logger.info(f"[{session_id}] Excel to FASTA conversion completed: {output_file_path} ({num_sequences} sequences, {file_size:.2f} KB)")
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": {
                "status": "success",
                "output_file": str(output_file_path.absolute()),
                "input_file": str(input_file_path.absolute()),
                "num_sequences": num_sequences,
                "file_size_kb": round(file_size, 2),
                "format": "FASTA",
                "conversion_type": "excel_to_fasta"
            }
        }
        
    except ImportError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower():
            error_msg = f"[{session_id}] Excel file detected but required dependency 'openpyxl' is missing. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] Missing required dependency: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "missing_dependency",
                "message": error_msg
            }
        }
    except ValueError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] Excel file detected but required dependency 'openpyxl' is missing. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] Excel to FASTA conversion failed (ValueError): {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "conversion_failed",
                "message": error_msg
            }
        }
    except Exception as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] Excel file detected but required dependency 'openpyxl' is missing. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] Unexpected error during Excel to FASTA conversion: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "unknown",
                "message": error_msg
            }
        }
    finally:
        # 清理从 URL 下载的临时输入文件
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.debug(f"[{session_id}] Cleaned up temporary input file: {temp_file_path}")
            except:
                pass


@mcp.tool()
async def convert_xlsx_to_fasta(args: ConvertXlsxToFastaArgs):
    """
    将 XLSX 文件转换为 FASTA 格式
    
    自动检测列名，支持多种常见的列名变体。
    支持 XLSX 格式（Excel 2007+）。
    支持本地文件路径和 HTTP/HTTPS URL。
    
    通过 SSE 流式推送转换进度，支持大文件转换。
    
    Args:
        args: ConvertXlsxToFastaArgs - 转换参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting XLSX to FASTA conversion")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting XLSX to FASTA conversion",
            "timestamp": time.time()
        }
    }
    
    temp_file_path = None
    input_file_path = None
    
    try:
        # 处理输入文件：可能是 URL 或本地路径
        if args.input_file.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Input is URL, downloading first: {args.input_file}")
            yield {
                "type": "progress",
                "data": {
                    "session_id": session_id,
                    "status": "downloading",
                    "message": f"Downloading input file from {args.input_file}",
                    "timestamp": time.time()
                }
            }
            try:
                temp_file_path = download_to_temp_file(args.input_file)
                input_file_path = Path(temp_file_path)
                logger.info(f"[{session_id}] Downloaded input file to: {input_file_path}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download input file from URL {args.input_file}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            input_file_path = Path(args.input_file)
            if not input_file_path.exists():
                error_msg = f"[{session_id}] Input file not found: {args.input_file}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
        else:
            # 自动生成输出文件名：基于输入文件名，替换扩展名为 .fasta
            output_file_path = input_file_path.with_suffix('.fasta')
            logger.info(f"[{session_id}] Auto-generated output path: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 执行转换，指定期望格式为 XLSX
        logger.info(f"[{session_id}] Converting XLSX to FASTA: {input_file_path} -> {output_file_path}")
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "converting",
                "message": f"Converting XLSX to FASTA: {input_file_path.name}",
                "timestamp": time.time()
            }
        }
        _convert_excel_to_fasta(input_file_path, output_file_path, expected_format='.xlsx')
        
        # 验证输出文件
        if not output_file_path.exists():
            error_msg = f"[{session_id}] Output FASTA file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "conversion_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size / 1024  # KB
        
        # 统计序列数量
        num_sequences = 0
        with open(output_file_path, 'r') as f:
            num_sequences = sum(1 for line in f if line.startswith('>'))
        
        logger.info(f"[{session_id}] XLSX to FASTA conversion completed: {output_file_path} ({num_sequences} sequences, {file_size:.2f} KB)")
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": {
                "status": "success",
                "output_file": str(output_file_path.absolute()),
                "input_file": str(input_file_path.absolute()),
                "num_sequences": num_sequences,
                "file_size_kb": round(file_size, 2),
                "format": "FASTA",
                "conversion_type": "xlsx_to_fasta"
            }
        }
        
    except ImportError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower():
            error_msg = f"[{session_id}] XLSX file detected but required dependency 'openpyxl' is missing. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] Missing required dependency: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "missing_dependency",
                "message": error_msg
            }
        }
    except ValueError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLSX file detected but required dependency 'openpyxl' is missing. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] XLSX to FASTA conversion failed (ValueError): {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "conversion_failed",
                "message": error_msg
            }
        }
    except Exception as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLSX file detected but required dependency 'openpyxl' is missing. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] Unexpected error during XLSX to FASTA conversion: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "unknown",
                "message": error_msg
            }
        }
    finally:
        # 清理从 URL 下载的临时输入文件
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.debug(f"[{session_id}] Cleaned up temporary input file: {temp_file_path}")
            except:
                pass


@mcp.tool()
async def convert_xls_to_fasta(args: ConvertXlsToFastaArgs):
    """
    将 XLS 文件转换为 FASTA 格式
    
    自动检测列名，支持多种常见的列名变体。
    支持 XLS 格式（Excel 97-2003）。
    支持本地文件路径和 HTTP/HTTPS URL。
    
    通过 SSE 流式推送转换进度，支持大文件转换。
    
    Args:
        args: ConvertXlsToFastaArgs - 转换参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting XLS to FASTA conversion")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting XLS to FASTA conversion",
            "timestamp": time.time()
        }
    }
    
    temp_file_path = None
    input_file_path = None
    
    try:
        # 处理输入文件：可能是 URL 或本地路径
        if args.input_file.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Input is URL, downloading first: {args.input_file}")
            yield {
                "type": "progress",
                "data": {
                    "session_id": session_id,
                    "status": "downloading",
                    "message": f"Downloading input file from {args.input_file}",
                    "timestamp": time.time()
                }
            }
            try:
                temp_file_path = download_to_temp_file(args.input_file)
                input_file_path = Path(temp_file_path)
                logger.info(f"[{session_id}] Downloaded input file to: {input_file_path}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download input file from URL {args.input_file}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            input_file_path = Path(args.input_file)
            if not input_file_path.exists():
                error_msg = f"[{session_id}] Input file not found: {args.input_file}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
        else:
            # 自动生成输出文件名：基于输入文件名，替换扩展名为 .fasta
            output_file_path = input_file_path.with_suffix('.fasta')
            logger.info(f"[{session_id}] Auto-generated output path: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 执行转换，指定期望格式为 XLS
        logger.info(f"[{session_id}] Converting XLS to FASTA: {input_file_path} -> {output_file_path}")
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "converting",
                "message": f"Converting XLS to FASTA: {input_file_path.name}",
                "timestamp": time.time()
            }
        }
        _convert_excel_to_fasta(input_file_path, output_file_path, expected_format='.xls')
        
        # 验证输出文件
        if not output_file_path.exists():
            error_msg = f"[{session_id}] Output FASTA file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "conversion_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size / 1024  # KB
        
        # 统计序列数量
        num_sequences = 0
        with open(output_file_path, 'r') as f:
            num_sequences = sum(1 for line in f if line.startswith('>'))
        
        logger.info(f"[{session_id}] XLS to FASTA conversion completed: {output_file_path} ({num_sequences} sequences, {file_size:.2f} KB)")
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": {
                "status": "success",
                "output_file": str(output_file_path.absolute()),
                "input_file": str(input_file_path.absolute()),
                "num_sequences": num_sequences,
                "file_size_kb": round(file_size, 2),
                "format": "FASTA",
                "conversion_type": "xls_to_fasta"
            }
        }
        
    except ImportError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'xlrd' in error_str.lower():
            error_msg = f"[{session_id}] XLS file detected but required dependency is missing. " \
                       f"Please install it using: pip install xlrd or conda install xlrd"
        else:
            error_msg = f"[{session_id}] Missing required dependency: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "missing_dependency",
                "message": error_msg
            }
        }
    except ValueError as e:
        error_str = str(e)
        if 'xlrd' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLS file detected but required dependency 'xlrd' is missing. " \
                       f"Please install it using: pip install xlrd or conda install xlrd"
        else:
            error_msg = f"[{session_id}] XLS to FASTA conversion failed (ValueError): {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "conversion_failed",
                "message": error_msg
            }
        }
    except Exception as e:
        error_str = str(e)
        if 'xlrd' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLS file detected but required dependency 'xlrd' is missing. " \
                       f"Please install it using: pip install xlrd or conda install xlrd"
        else:
            error_msg = f"[{session_id}] Unexpected error during XLS to FASTA conversion: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "unknown",
                "message": error_msg
            }
        }
    finally:
        # 清理从 URL 下载的临时输入文件
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.debug(f"[{session_id}] Cleaned up temporary input file: {temp_file_path}")
            except:
                pass


@mcp.tool()
async def convert_xlsx_to_csv(args: ConvertXlsxToCsvArgs):
    """
    将 XLSX 文件转换为 CSV 格式
    
    支持 XLSX 格式（Excel 2007+）。
    保留所有列和数据，不进行任何过滤。
    支持本地文件路径和 HTTP/HTTPS URL。
    
    通过 SSE 流式推送转换进度，支持大文件转换。
    
    Args:
        args: ConvertXlsxToCsvArgs - 转换参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting XLSX to CSV conversion")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting XLSX to CSV conversion",
            "timestamp": time.time()
        }
    }
    
    temp_file_path = None
    input_file_path = None
    
    try:
        # 处理输入文件：可能是 URL 或本地路径
        if args.input_file.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Input is URL, downloading first: {args.input_file}")
            yield {
                "type": "progress",
                "data": {
                    "session_id": session_id,
                    "status": "downloading",
                    "message": f"Downloading input file from {args.input_file}",
                    "timestamp": time.time()
                }
            }
            try:
                temp_file_path = download_to_temp_file(args.input_file)
                input_file_path = Path(temp_file_path)
                logger.info(f"[{session_id}] Downloaded input file to: {input_file_path}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download input file from URL {args.input_file}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            input_file_path = Path(args.input_file)
            if not input_file_path.exists():
                error_msg = f"[{session_id}] Input file not found: {args.input_file}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
        else:
            # 自动生成输出文件名：基于输入文件名，替换扩展名为 .csv
            output_file_path = input_file_path.with_suffix('.csv')
            logger.info(f"[{session_id}] Auto-generated output path: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 执行转换，指定期望格式为 XLSX
        logger.info(f"[{session_id}] Converting XLSX to CSV: {input_file_path} -> {output_file_path}")
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "converting",
                "message": f"Converting XLSX to CSV: {input_file_path.name}",
                "timestamp": time.time()
            }
        }
        _convert_excel_to_csv(input_file_path, output_file_path, expected_format='.xlsx')
        
        # 验证输出文件
        if not output_file_path.exists():
            error_msg = f"[{session_id}] Output CSV file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "conversion_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size / 1024  # KB
        
        # 读取结果文件统计行数和列数
        result_df = pd.read_csv(output_file_path)
        num_rows = len(result_df)
        num_columns = len(result_df.columns)
        
        logger.info(f"[{session_id}] XLSX to CSV conversion completed: {output_file_path} ({num_rows} rows, {num_columns} columns, {file_size:.2f} KB)")
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": {
                "status": "success",
                "output_file": str(output_file_path.absolute()),
                "input_file": str(input_file_path.absolute()),
            "num_rows": num_rows,
            "num_columns": num_columns,
            "file_size_kb": round(file_size, 2),
            "format": "CSV",
            "conversion_type": "xlsx_to_csv"
            }
        }
        
    except ImportError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower():
            error_msg = f"[{session_id}] XLSX file detected but required dependency 'openpyxl' is missing. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] Missing required dependency: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "missing_dependency",
                "message": error_msg
            }
        }
    except ValueError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLSX file detected but required dependency 'openpyxl' is missing. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] XLSX to CSV conversion failed (ValueError): {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "conversion_failed",
                "message": error_msg
            }
        }
    except Exception as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLSX file detected but required dependency 'openpyxl' is missing. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] Unexpected error during XLSX to CSV conversion: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "unknown",
                "message": error_msg
            }
        }
    finally:
        # 清理从 URL 下载的临时输入文件
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.debug(f"[{session_id}] Cleaned up temporary input file: {temp_file_path}")
            except:
                pass


@mcp.tool()
async def convert_xls_to_csv(args: ConvertXlsToCsvArgs):
    """
    将 XLS 文件转换为 CSV 格式
    
    支持 XLS 格式（Excel 97-2003）。
    保留所有列和数据，不进行任何过滤。
    支持本地文件路径和 HTTP/HTTPS URL。
    
    通过 SSE 流式推送转换进度，支持大文件转换。
    
    Args:
        args: ConvertXlsToCsvArgs - 转换参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting XLS to CSV conversion")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting XLS to CSV conversion",
            "timestamp": time.time()
        }
    }
    
    temp_file_path = None
    input_file_path = None
    
    try:
        # 处理输入文件：可能是 URL 或本地路径
        if args.input_file.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Input is URL, downloading first: {args.input_file}")
            try:
                temp_file_path = download_to_temp_file(args.input_file)
                input_file_path = Path(temp_file_path)
                logger.info(f"[{session_id}] Downloaded input file to: {input_file_path}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download input file from URL {args.input_file}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            input_file_path = Path(args.input_file)
            if not input_file_path.exists():
                error_msg = f"[{session_id}] Input file not found: {args.input_file}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
        else:
            # 自动生成输出文件名：基于输入文件名，替换扩展名为 .csv
            output_file_path = input_file_path.with_suffix('.csv')
            logger.info(f"[{session_id}] Auto-generated output path: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 执行转换，指定期望格式为 XLS
        logger.info(f"[{session_id}] Converting XLS to CSV: {input_file_path} -> {output_file_path}")
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "converting",
                "message": f"Converting XLS to CSV: {input_file_path.name}",
                "timestamp": time.time()
            }
        }
        _convert_excel_to_csv(input_file_path, output_file_path, expected_format='.xls')
        
        # 验证输出文件
        if not output_file_path.exists():
            error_msg = f"[{session_id}] Output CSV file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "conversion_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size / 1024  # KB
        
        # 读取结果文件统计行数和列数
        result_df = pd.read_csv(output_file_path)
        num_rows = len(result_df)
        num_columns = len(result_df.columns)
        
        logger.info(f"[{session_id}] XLS to CSV conversion completed: {output_file_path} ({num_rows} rows, {num_columns} columns, {file_size:.2f} KB)")
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": {
                "status": "success",
                "output_file": str(output_file_path.absolute()),
                "input_file": str(input_file_path.absolute()),
                "num_rows": num_rows,
                "num_columns": num_columns,
                "file_size_kb": round(file_size, 2),
                "format": "CSV",
                "conversion_type": "xls_to_csv"
            }
        }
        
    except ImportError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'xlrd' in error_str.lower():
            error_msg = f"[{session_id}] XLS file detected but required dependency is missing. " \
                       f"Please install it using: pip install openpyxl xlrd or conda install openpyxl xlrd"
        else:
            error_msg = f"[{session_id}] Missing required dependency: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "missing_dependency",
                "message": error_msg
            }
        }
    except ValueError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'xlrd' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLS file detected but required dependency is missing. " \
                       f"Please install it using: pip install openpyxl xlrd or conda install openpyxl xlrd"
        else:
            error_msg = f"[{session_id}] XLS to CSV conversion failed (ValueError): {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "conversion_failed",
                "message": error_msg
            }
        }
    except Exception as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'xlrd' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLS file detected but required dependency is missing. " \
                       f"Please install it using: pip install openpyxl xlrd or conda install openpyxl xlrd"
        else:
            error_msg = f"[{session_id}] Unexpected error during XLS to CSV conversion: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "unknown",
                "message": error_msg
            }
        }
    finally:
        # 清理从 URL 下载的临时输入文件
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.debug(f"[{session_id}] Cleaned up temporary input file: {temp_file_path}")
            except:
                pass


@mcp.tool()
async def convert_csv_to_xlsx(args: ConvertCsvToXlsxArgs):
    """
    将 CSV 文件转换为 XLSX 格式
    
    支持 CSV 格式输入。
    保留所有列和数据，不进行任何过滤。
    支持本地文件路径和 HTTP/HTTPS URL。
    
    Args:
        args: ConvertCsvToXlsxArgs - 转换参数
        
    Returns:
        包含输出文件路径和转换信息的字典
    """
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting CSV to XLSX conversion")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting CSV to XLSX conversion",
            "timestamp": time.time()
        }
    }
    
    temp_file_path = None
    input_file_path = None
    
    try:
        # 处理输入文件：可能是 URL 或本地路径
        if args.input_file.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Input is URL, downloading first: {args.input_file}")
            try:
                temp_file_path = download_to_temp_file(args.input_file)
                input_file_path = Path(temp_file_path)
                logger.info(f"[{session_id}] Downloaded input file to: {input_file_path}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download input file from URL {args.input_file}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            input_file_path = Path(args.input_file)
            if not input_file_path.exists():
                error_msg = f"[{session_id}] Input file not found: {args.input_file}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
        else:
            # 自动生成输出文件名：基于输入文件名，替换扩展名为 .xlsx
            output_file_path = input_file_path.with_suffix('.xlsx')
            logger.info(f"[{session_id}] Auto-generated output path: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 执行转换
        logger.info(f"[{session_id}] Converting CSV to XLSX: {input_file_path} -> {output_file_path}")
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "converting",
                "message": f"Converting CSV to XLSX: {input_file_path.name}",
                "timestamp": time.time()
            }
        }
        _convert_csv_to_excel(input_file_path, output_file_path, output_format='xlsx')
        
        # 验证输出文件
        if not output_file_path.exists():
            error_msg = f"[{session_id}] Output XLSX file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "conversion_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size / 1024  # KB
        
        # 读取结果文件统计行数和列数
        result_df = pd.read_excel(output_file_path)
        num_rows = len(result_df)
        num_columns = len(result_df.columns)
        
        logger.info(f"[{session_id}] CSV to XLSX conversion completed: {output_file_path} ({num_rows} rows, {num_columns} columns, {file_size:.2f} KB)")
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": {
                "status": "success",
                "output_file": str(output_file_path.absolute()),
                "input_file": str(input_file_path.absolute()),
                "num_rows": num_rows,
                "num_columns": num_columns,
                "file_size_kb": round(file_size, 2),
                "format": "XLSX",
                "conversion_type": "csv_to_xlsx"
            }
        }
        
    except ImportError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower():
            error_msg = f"[{session_id}] XLSX file writing requires dependency 'openpyxl'. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] Missing required dependency: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "missing_dependency",
                "message": error_msg
            }
        }
    except ValueError as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLSX file writing requires dependency 'openpyxl'. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] CSV to XLSX conversion failed (ValueError): {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "conversion_failed",
                "message": error_msg
            }
        }
    except Exception as e:
        error_str = str(e)
        if 'openpyxl' in error_str.lower() or 'Missing optional dependency' in error_str:
            error_msg = f"[{session_id}] XLSX file writing requires dependency 'openpyxl'. " \
                       f"Please install it using: pip install openpyxl or conda install openpyxl"
        else:
            error_msg = f"[{session_id}] Unexpected error during CSV to XLSX conversion: {error_str}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "unknown",
                "message": error_msg
            }
        }
    finally:
        # 清理从 URL 下载的临时输入文件
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
                logger.debug(f"[{session_id}] Cleaned up temporary input file: {temp_file_path}")
            except:
                pass


@mcp.tool()
async def merge_csv_cartesian(args: MergeCsvCartesianArgs):
    """
    将抗体 CSV 文件和抗原 CSV 文件进行笛卡尔积合并
    
    抗体文件的每一行都与抗原文件的每一行组合，生成所有可能的抗体-抗原对。
    结果行数 = 抗体文件行数 × 抗原文件行数
    两个文件的表头都会保留在结果中，如果列名冲突，会自动添加后缀。
    支持本地文件路径和 HTTP/HTTPS URL。
    
    抗体文件必须包含 'Heavy' 和 'Light' 列。
    抗原文件必须包含 'variant_seq'（或 'antigen_seq'、'variant'）列。
    
    通过 SSE 流式推送合并进度，支持大文件合并。
    
    Args:
        args: MergeCsvCartesianArgs - 合并参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting antibody-antigen CSV cartesian merge")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting CSV cartesian merge",
            "timestamp": time.time()
        }
    }
    
    temp_file_path_ab = None
    temp_file_path_ag = None
    antibody_file_path = None
    antigen_file_path = None
    
    try:
        # 处理抗体文件：可能是 URL 或本地路径
        if args.antibody_file.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Antibody file is URL, downloading first: {args.antibody_file}")
            try:
                temp_file_path_ab = download_to_temp_file(args.antibody_file)
                antibody_file_path = Path(temp_file_path_ab)
                logger.info(f"[{session_id}] Downloaded antibody file to: {antibody_file_path}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download antibody file from URL {args.antibody_file}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            antibody_file_path = Path(args.antibody_file)
            if not antibody_file_path.exists():
                error_msg = f"[{session_id}] Antibody file not found: {args.antibody_file}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 处理抗原文件：可能是 URL 或本地路径
        if args.antigen_file.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Antigen file is URL, downloading first: {args.antigen_file}")
            try:
                temp_file_path_ag = download_to_temp_file(args.antigen_file)
                antigen_file_path = Path(temp_file_path_ag)
                logger.info(f"[{session_id}] Downloaded antigen file to: {antigen_file_path}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download antigen file from URL {args.antigen_file}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            antigen_file_path = Path(args.antigen_file)
            if not antigen_file_path.exists():
                error_msg = f"[{session_id}] Antigen file not found: {args.antigen_file}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
        else:
            # 自动生成随机输出文件名（避免文件名过长）
            random_id = str(uuid.uuid4())[:8]
            output_file_path = antibody_file_path.parent / f"merged_{random_id}.csv"
            logger.info(f"[{session_id}] Auto-generated output path: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 执行合并
        logger.info(f"[{session_id}] Merging antibody and antigen CSV files (cartesian): {antibody_file_path} × {antigen_file_path} -> {output_file_path}")
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "merging",
                "message": f"Merging CSV files (cartesian): {antibody_file_path.name} × {antigen_file_path.name}",
                "timestamp": time.time()
            }
        }
        
        # 检查是否提供了 key 列
        if args.antibody_key and args.antigen_key:
            logger.info(f"[{session_id}] Will create new key column '{args.new_key_name}' from '{args.antibody_key}' (antibody) and '{args.antigen_key}' (antigen)")
            _merge_csv_cartesian(
                antibody_file_path, 
                antigen_file_path, 
                output_file_path,
                key_column1=args.antibody_key,
                key_column2=args.antigen_key,
                new_key_name=args.new_key_name
            )
        elif args.key_column1 or args.key_column2:
            # 只提供了一个 key 列，给出警告
            logger.warning(f"[{session_id}] Only one key column provided. Both antibody_key and antigen_key are required to create new key. Skipping key generation.")
            _merge_csv_cartesian(antibody_file_path, antigen_file_path, output_file_path)
        else:
            _merge_csv_cartesian(antibody_file_path, antigen_file_path, output_file_path)
        
        # 验证输出文件
        if not output_file_path.exists():
            error_msg = f"[{session_id}] Merged CSV file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "merge_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size / 1024  # KB
        
        # 读取结果文件统计行数
        result_df = pd.read_csv(output_file_path)
        num_rows = len(result_df)
        
        logger.info(f"[{session_id}] Antibody-antigen CSV cartesian merge completed: {output_file_path} ({num_rows} rows, {file_size:.2f} KB)")
        
        result = {
            "status": "success",
            "output_file": str(output_file_path.absolute()),
            "antibody_file": str(antibody_file_path.absolute()),
            "antigen_file": str(antigen_file_path.absolute()),
            "num_rows": num_rows,
            "num_columns": len(result_df.columns),
            "file_size_kb": round(file_size, 2),
            "format": "CSV",
            "merge_type": "cartesian",
            "description": "Antibody-antigen cartesian merge"
        }
        
        # 如果创建了新 key 列，添加到返回结果中
        if args.antibody_key and args.antigen_key:
            result["new_key_column"] = args.new_key_name
            result["antibody_key"] = args.antibody_key
            result["antigen_key"] = args.antigen_key
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": result
        }
        
    except ValueError as e:
        error_msg = f"[{session_id}] Antibody-antigen CSV cartesian merge failed (ValueError): {str(e)}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "merge_failed",
                "message": error_msg
            }
        }
    except Exception as e:
        error_msg = f"[{session_id}] Unexpected error during antibody-antigen CSV cartesian merge: {str(e)}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "unknown",
                "message": error_msg
            }
        }
    finally:
        # 清理从 URL 下载的临时输入文件
        for temp_file_path in [temp_file_path_ab, temp_file_path_ag]:
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                    logger.debug(f"[{session_id}] Cleaned up temporary input file: {temp_file_path}")
                except:
                    pass


@mcp.tool()
async def merge_csv_by_key(args: MergeCsvByKeyArgs):
    """
    合并两个 CSV 文件（加法运算）
    
    如果提供了 key_column，则基于指定 key 列合并，要求两个文件行数一致，key 值一一对应。
    如果没有提供 key_column，则按照行索引一一对应进行拼接。
    表头取并集，如果列名冲突（除了 key 列），会自动添加后缀。
    支持本地文件路径和 HTTP/HTTPS URL。
    
    通过 SSE 流式推送合并进度，支持大文件合并。
    
    Args:
        args: MergeCsvByKeyArgs - 合并参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    session_id = str(uuid.uuid4())[:8]
    if args.key_column:
        logger.info(f"[{session_id}] Starting CSV key-based merge with key: {args.key_column}")
    else:
        logger.info(f"[{session_id}] Starting CSV row-index-based merge")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting CSV merge",
            "timestamp": time.time()
        }
    }
    
    temp_file_path1 = None
    temp_file_path2 = None
    input_file_path1 = None
    input_file_path2 = None
    
    try:
        # 处理第一个输入文件：可能是 URL 或本地路径
        if args.input_file1.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Input file 1 is URL, downloading first: {args.input_file1}")
            try:
                temp_file_path1 = download_to_temp_file(args.input_file1)
                input_file_path1 = Path(temp_file_path1)
                logger.info(f"[{session_id}] Downloaded input file 1 to: {input_file_path1}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download input file 1 from URL {args.input_file1}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            input_file_path1 = Path(args.input_file1)
            if not input_file_path1.exists():
                error_msg = f"[{session_id}] Input file 1 not found: {args.input_file1}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 处理第二个输入文件：可能是 URL 或本地路径
        if args.input_file2.startswith(('http://', 'https://')):
            logger.info(f"[{session_id}] Input file 2 is URL, downloading first: {args.input_file2}")
            try:
                temp_file_path2 = download_to_temp_file(args.input_file2)
                input_file_path2 = Path(temp_file_path2)
                logger.info(f"[{session_id}] Downloaded input file 2 to: {input_file_path2}")
            except Exception as e:
                error_msg = f"[{session_id}] Failed to download input file 2 from URL {args.input_file2}: {str(e)}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "download_failed",
                        "message": error_msg
                    }
                }
                return
        else:
            input_file_path2 = Path(args.input_file2)
            if not input_file_path2.exists():
                error_msg = f"[{session_id}] Input file 2 not found: {args.input_file2}"
                logger.error(error_msg)
                yield {
                    "type": "error",
                    "data": {
                        "session_id": session_id,
                        "error_type": "file_not_found",
                        "message": error_msg
                    }
                }
                return
        
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
        else:
            # 自动生成随机输出文件名（避免文件名过长）
            random_id = str(uuid.uuid4())[:8]
            output_file_path = input_file_path1.parent / f"merged_{random_id}.csv"
            logger.info(f"[{session_id}] Auto-generated output path: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 执行合并
        if args.key_column:
            logger.info(f"[{session_id}] Merging CSV files by key '{args.key_column}': {input_file_path1} + {input_file_path2} -> {output_file_path}")
        else:
            logger.info(f"[{session_id}] Merging CSV files by row index: {input_file_path1} + {input_file_path2} -> {output_file_path}")
        
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "merging",
                "message": f"Merging CSV files: {input_file_path1.name} + {input_file_path2.name}",
                "timestamp": time.time()
            }
        }
        
        try:
            _merge_csv_by_key(input_file_path1, input_file_path2, output_file_path, key_column=args.key_column)
        except Exception as e:
            # 如果合并失败，检查是否是因为其中一个文件为空
            # 这种情况下，_merge_csv_by_key 应该已经处理了，但如果还有错误，我们需要处理
            error_msg = f"[{session_id}] CSV merge failed: {str(e)}"
            logger.error(error_msg, exc_info=True)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "merge_failed",
                    "message": error_msg
                }
            }
        
        # 验证输出文件
        if not output_file_path.exists():
            error_msg = f"[{session_id}] Merged CSV file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "merge_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size / 1024  # KB
        
        # 读取结果文件统计行数（如果文件不为空）
        num_rows = 0
        num_columns = 0
        if file_size > 0:
            try:
                result_df = pd.read_csv(output_file_path)
                num_rows = len(result_df)
                num_columns = len(result_df.columns)
            except Exception as e:
                logger.warning(f"[{session_id}] Could not read output file for statistics: {str(e)}")
        
        merge_type = "key_based" if args.key_column else "row_index"
        merge_desc = f"key '{args.key_column}'" if args.key_column else "row index"
        logger.info(f"[{session_id}] CSV merge by {merge_desc} completed: {output_file_path} ({num_rows} rows, {num_columns} columns, {file_size:.2f} KB)")
        
        result = {
            "status": "success",
            "output_file": str(output_file_path.absolute()),
            "input_file1": str(input_file_path1.absolute()),
            "input_file2": str(input_file_path2.absolute()),
            "num_rows": num_rows,
            "num_columns": num_columns,
            "file_size_kb": round(file_size, 2),
            "format": "CSV",
            "merge_type": merge_type
        }
        
        # 如果提供了 key_column，添加到返回结果中
        if args.key_column:
            result["key_column"] = args.key_column
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": result
        }
        
    except ValueError as e:
        error_msg = f"[{session_id}] CSV key-based merge failed (ValueError): {str(e)}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "merge_failed",
                "message": error_msg
            }
        }
    except Exception as e:
        error_msg = f"[{session_id}] Unexpected error during CSV key-based merge: {str(e)}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "unknown",
                "message": error_msg
            }
        }
    finally:
        # 清理从 URL 下载的临时输入文件
        for temp_file_path in [temp_file_path1, temp_file_path2]:
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                    logger.debug(f"[{session_id}] Cleaned up temporary input file: {temp_file_path}")
                except:
                    pass


@mcp.tool()
async def create_csv(args: CreateCsvArgs):
    """
    创建一个空白的 CSV 文件
    
    可以创建完全空白的 CSV 文件，或者包含指定列名的表头。
    如果未指定输出路径，将创建临时文件。
    
    通过 SSE 流式推送创建进度。
    
    Args:
        args: CreateCsvArgs - 创建参数
        
    Yields:
        Progress updates and final result through SSE stream.
    """
    session_id = str(uuid.uuid4())[:8]
    logger.info(f"[{session_id}] Starting CSV file creation")
    
    # 发送初始化进度
    yield {
        "type": "progress",
        "data": {
            "session_id": session_id,
            "status": "initializing",
            "message": "Starting CSV file creation",
            "timestamp": time.time()
        }
    }
    
    try:
        # 确定输出文件路径
        if args.output_file:
            output_file_path = Path(args.output_file)
            logger.info(f"[{session_id}] Using specified output path: {output_file_path}")
        else:
            # 创建临时文件
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv')
            output_file_path = Path(temp_file.name)
            temp_file.close()
            logger.info(f"[{session_id}] Created temporary CSV file: {output_file_path}")
        
        # 确保输出目录存在
        output_file_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 创建 CSV 文件
        logger.debug(f"[{session_id}] Writing CSV file: {output_file_path}")
        yield {
            "type": "progress",
            "data": {
                "session_id": session_id,
                "status": "creating",
                "message": f"Creating CSV file: {output_file_path.name}",
                "timestamp": time.time()
            }
        }
        
        if args.columns:
            # 如果提供了列名，写入表头
            # 处理列名：可能是逗号分隔的字符串或列表
            if isinstance(args.columns, str):
                # 如果是字符串，尝试分割
                column_list = [col.strip() for col in args.columns.split(',')]
            elif isinstance(args.columns, list):
                column_list = args.columns
            else:
                column_list = []
            
            if column_list:
                # 创建包含表头的 CSV 文件
                df = pd.DataFrame(columns=column_list)
                df.to_csv(output_file_path, index=False, encoding='utf-8')
                logger.info(f"[{session_id}] Created CSV file with {len(column_list)} columns: {', '.join(column_list)}")
            else:
                # 列名为空，创建空白文件
                with open(output_file_path, 'w', encoding='utf-8') as f:
                    pass
                logger.info(f"[{session_id}] Created empty CSV file")
        else:
            # 没有提供列名，创建完全空白的 CSV 文件
            with open(output_file_path, 'w', encoding='utf-8') as f:
                pass
            logger.info(f"[{session_id}] Created blank CSV file")
        
        # 验证文件是否成功创建
        if not output_file_path.exists():
            error_msg = f"[{session_id}] CSV file was not created: {output_file_path}"
            logger.error(error_msg)
            yield {
                "type": "error",
                "data": {
                    "session_id": session_id,
                    "error_type": "file_creation_failed",
                    "message": error_msg
                }
            }
            return
        
        file_size = output_file_path.stat().st_size  # bytes
        is_temporary = args.output_file is None
        
        logger.info(f"[{session_id}] CSV file created successfully: {output_file_path} ({file_size} bytes)")
        
        result = {
            "status": "success",
            "file_path": str(output_file_path.absolute()),
            "file_size_bytes": file_size,
            "is_temporary": is_temporary
        }
        
        # 如果提供了列名，添加到返回结果中
        if args.columns:
            if isinstance(args.columns, str):
                column_list = [col.strip() for col in args.columns.split(',')]
            elif isinstance(args.columns, list):
                column_list = args.columns
            else:
                column_list = []
            result["columns"] = column_list
            result["num_columns"] = len(column_list)
        else:
            result["num_columns"] = 0
        
        # 返回最终结果
        yield {
            "type": "result",
            "data": result
        }
        
    except Exception as e:
        error_msg = f"[{session_id}] Failed to create CSV file: {str(e)}"
        logger.error(error_msg, exc_info=True)
        yield {
            "type": "error",
            "data": {
                "session_id": session_id,
                "error_type": "creation_failed",
                "message": error_msg
            }
        }


# 添加生命周期管理
@asynccontextmanager
async def file_utils_lifespan(server: FastMCP) -> AsyncIterator[dict]:
    """处理服务器启动和关闭"""
    print("File Utils MCP Server 正在初始化...")
    
    try:
        yield {"initialized": True}
    finally:
        print("File Utils MCP Server 正在关闭...")

# 设置生命周期
mcp.lifespan = file_utils_lifespan

if __name__ == "__main__":
    print("启动 File Utils MCP 服务器...")
    # 设置网络参数
    mcp.settings.host = "0.0.0.0"
    mcp.settings.port = 8110  # 使用不同的端口
    
    # 使用SSE模式启动
    mcp.run(transport="sse")

